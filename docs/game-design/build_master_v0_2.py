#!/usr/bin/env python3
"""Build WildLive-Game-Master-Draft-v0.2.xlsx.

v0.2 revises the v0.1 draft along the lines the human review of v0.1
decided (see the v0.2 Review sheet for the paper trail):

    * Africa-first — the initial playable roster is Africa-only.
      Non-African maps are still in the workbook but tagged
      `availability_phase = future_expansion`; the game unlocks them
      as a future release, not on day 1.
    * Real habitats — every MapAnimals row respects the species'
      real biogeographic range. If a species has no home in v0.2's
      map set, it is tagged `availability_phase = future_region` and
      has NO MapAnimals row until the required expansion map ships.
    * Northern White Rhinoceros removed from normal spawn — tagged
      `availability_phase = special_event` and has NO MapAnimals row.
    * Hunter is not owned — `hire_cost_g` renamed to `contract_cost_g`
      to make the one-off Guild-pool contract model explicit.
    * Map unlock is separated from Hunter rank — new columns
      `unlock_rule` / `unlock_value` / `recommended_hunter_rank`
      replace v0.1's single `unlock_rank`. Low-rank Hunters can still
      accept high-rank maps (higher failure rate) — access permitted,
      risk high.
    * Biomes master — the free-text `biome` / `habitat` /
      `preferred_biome` columns from v0.1 now reference a shared
      Biomes sheet with FK validation.
    * A few tightly-associated real-world names in v0.1 (Tenzing
      Sherpa in particular) are replaced with clearly fictional
      names.
    * Two new specialist African maps — Ethiopian Highlands and
      Virunga Highlands — so Ethiopian Wolf and Mountain Gorilla can
      live in their real ranges.

v0.1 files under docs/game-design/ are NOT modified.

Run:

    python3 docs/game-design/build_master_v0_2.py

Output:

    docs/game-design/WildLive-Game-Master-Draft-v0.2.xlsx
"""

from __future__ import annotations

import io
import re
import sys
import zipfile
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Byte-deterministic XLSX save.
#
# openpyxl stamps the current wall-clock into docProps/core.xml and into the
# ZIP entry timestamps at save time. That means the same source produces
# different bytes on every run and `git diff` is never clean.
#
# _deterministic_save() normalises the three run-varying pieces:
#   - dcterms:created / dcterms:modified in docProps/core.xml
#   - ZIP entry date_time on every part
#   - ZIP entry order (sorted alphabetically)
#
# Game data (cell values, formulas, IDs, FK links) is not touched — only the
# packaging metadata is canonicalised.
# ---------------------------------------------------------------------------

_CANONICAL_ISO = "2000-01-01T00:00:00Z"
_CANONICAL_ZIP_TUPLE = (2000, 1, 1, 0, 0, 0)
_CANONICAL_CREATOR = "wildlive-master-builder"


def _deterministic_save(wb: Workbook, out_path: Path) -> None:
    wb.properties.creator = _CANONICAL_CREATOR
    wb.properties.lastModifiedBy = _CANONICAL_CREATOR

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    with zipfile.ZipFile(buf, "r") as src:
        entries = sorted(src.namelist())
        payloads: List[Tuple[str, bytes]] = []
        for name in entries:
            data = src.read(name)
            if name == "docProps/core.xml":
                text = data.decode("utf-8")
                text = re.sub(
                    r"(<dcterms:created[^>]*>)[^<]*(</dcterms:created>)",
                    lambda m: m.group(1) + _CANONICAL_ISO + m.group(2),
                    text,
                )
                text = re.sub(
                    r"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                    lambda m: m.group(1) + _CANONICAL_ISO + m.group(2),
                    text,
                )
                data = text.encode("utf-8")
            payloads.append((name, data))

    with open(out_path, "wb") as f:
        with zipfile.ZipFile(f, "w", zipfile.ZIP_DEFLATED) as dst:
            for name, data in payloads:
                zi = zipfile.ZipInfo(filename=name, date_time=_CANONICAL_ZIP_TUPLE)
                zi.compress_type = zipfile.ZIP_DEFLATED
                zi.external_attr = 0o644 << 16
                dst.writestr(zi, data)


# =============================================================================
# 1. Biomes (NEW in v0.2 — shared master)
#
# 6 biomes. Consolidated deliberately; more granular biomes (e.g. montane
# forest, mangrove) can be added when a Map or Animal actually needs them.
# =============================================================================

BIOMES: List[Dict[str, Any]] = [
    dict(biome_id="biome_savanna",    name_ja="サバンナ",
         name_en="Savanna",
         description_ja="疎らな木々を含む広大な熱帯草原。東・南アフリカに広がる。",
         description_en="Open tropical grassland with scattered trees. East and Southern Africa."),
    dict(biome_id="biome_wetland",    name_ja="湿地",
         name_en="Wetland",
         description_ja="季節性・恒久性の淡水氾濫原、沼沢、河川デルタ。",
         description_en="Freshwater floodplains, marshes, and river deltas."),
    dict(biome_id="biome_desert",     name_ja="砂漠",
         name_en="Desert",
         description_ja="極端に降水量の少ない乾燥地帯。サハラ、ナミブ、カラハリ、内陸オーストラリアなど。",
         description_en="Extremely arid low-rainfall zones. Sahara, Namib, Kalahari, Australian interior."),
    dict(biome_id="biome_mountain",   name_ja="山岳",
         name_en="Mountain",
         description_ja="高地の岩場・草地・雲霧林・高山斜面までを含む。",
         description_en="High-altitude terrain from rocky peaks to alpine slopes and montane forests."),
    dict(biome_id="biome_rainforest", name_ja="熱帯雨林",
         name_en="Rainforest",
         description_ja="低地から中高度の常緑広葉樹林。多層の林冠と豊富な降水量が特徴。",
         description_en="Lowland to mid-altitude evergreen broadleaf forest with multi-layer canopy and high rainfall."),
    dict(biome_id="biome_tundra",     name_ja="ツンドラ",
         name_en="Tundra",
         description_ja="極域の樹木限界以北の凍土平原。",
         description_en="Cold treeless plain of the polar regions above the tree line."),
]

BIOME_IDS = {b["biome_id"] for b in BIOMES}


# =============================================================================
# 2. Rarities  (5 tiers — unchanged from v0.1 for compatibility)
#
# Rarity definition is CLARIFIED in v0.2:
#   Rarity here = in-game encounter/capture rarity, NOT real-world
#   conservation status. A future `conservation_status` column may
#   record IUCN status separately (see Review).
# =============================================================================

RARITIES: List[Dict[str, Any]] = [
    dict(rarity_id="rarity_common",    name_en="Common",    name_ja="コモン",     sort_order=1, base_multiplier=1.0,
         description="In-game rarity: encountered on most easy expeditions. Not an IUCN status."),
    dict(rarity_id="rarity_uncommon",  name_en="Uncommon",  name_ja="アンコモン", sort_order=2, base_multiplier=1.5,
         description="In-game rarity: reliable mid-tier target."),
    dict(rarity_id="rarity_rare",      name_en="Rare",      name_ja="レア",       sort_order=3, base_multiplier=2.5,
         description="In-game rarity: iconic species most players will eventually collect."),
    dict(rarity_id="rarity_epic",      name_en="Epic",      name_ja="エピック",   sort_order=4, base_multiplier=5.0,
         description="In-game rarity: prestige tier. Real-world conservation status often but not always critical."),
    dict(rarity_id="rarity_legendary", name_en="Legendary", name_ja="レジェンド", sort_order=5, base_multiplier=12.0,
         description="In-game rarity: near-mythical encounter. Handful per player expected in a long play life."),
]
RARITY_IDS = {r["rarity_id"] for r in RARITIES}


# =============================================================================
# 3. Maps  (15 total — 9 initial African + 6 future expansion)
#
# New columns in v0.2:
#   availability_phase        initial_africa | future_expansion
#   unlock_rule               always | zoo_value | future_expansion
#   unlock_value              numeric threshold if unlock_rule = zoo_value
#   recommended_hunter_rank   guidance only (0..6); does NOT gate dispatch
#   minimum_hunter_rank_gate  hard gate (0 = none). Draft = 0 everywhere
#                              (access permitted, risk high — see Review)
#   biome_id                  FK → Biomes (replaces v0.1 free-text biome)
#
# Removed from v0.1:
#   unlock_rank               (conflated Map access with Hunter ownership;
#                               replaced by the separated columns above)
# =============================================================================

MAPS: List[Dict[str, Any]] = [
    # -- Initial Africa (9) ---------------------------------------------------
    dict(map_id="map_kenyan_savanna_001",
         name_en="Kenyan Savanna",       name_ja="ケニアのサバンナ",
         region="East Africa", biome_id="biome_savanna",
         availability_phase="initial_africa",
         unlock_rule="always", unlock_value=0,
         recommended_hunter_rank=1, minimum_hunter_rank_gate=0,
         difficulty=1, expedition_minutes=10, base_cost_g=50, risk_level=1,
         description_en="Wide-open grassland teeming with iconic African wildlife. First place every player starts.",
         description_ja="広大な草原に象徴的なアフリカの動物が集う地。すべてのプレイヤーが最初に訪れる場所。"),

    dict(map_id="map_serengeti_plains_002",
         name_en="Serengeti Plains",     name_ja="セレンゲティ平原",
         region="East Africa", biome_id="biome_savanna",
         availability_phase="initial_africa",
         unlock_rule="zoo_value", unlock_value=100,
         recommended_hunter_rank=1, minimum_hunter_rank_gate=0,
         difficulty=2, expedition_minutes=20, base_cost_g=90, risk_level=1,
         description_en="Endless migration corridor. Larger herds, but rangier terrain than a starter savanna.",
         description_ja="果てしない移動経路。群れは大きいが起伏に富み、初心者向けの草原よりやや荒々しい。"),

    dict(map_id="map_okavango_delta_003",
         name_en="Okavango Delta",       name_ja="オカバンゴ湿地",
         region="Southern Africa", biome_id="biome_wetland",
         availability_phase="initial_africa",
         unlock_rule="zoo_value", unlock_value=500,
         recommended_hunter_rank=2, minimum_hunter_rank_gate=0,
         difficulty=2, expedition_minutes=40, base_cost_g=180, risk_level=2,
         description_en="Seasonal wetland maze. Water-adapted species dominate; navigation is slow.",
         description_ja="季節性の湿地帯。水辺に適応した動物が主役で、移動は遅く難しい。"),

    dict(map_id="map_namib_desert_004",
         name_en="Namib Desert",         name_ja="ナミブ砂漠",
         region="Southern Africa", biome_id="biome_desert",
         availability_phase="initial_africa",
         unlock_rule="zoo_value", unlock_value=800,
         recommended_hunter_rank=2, minimum_hunter_rank_gate=0,
         difficulty=3, expedition_minutes=60, base_cost_g=240, risk_level=3,
         description_en="Extreme aridity. Few species, but the ones that live here are specialised and elusive.",
         description_ja="極度の乾燥地帯。生息種は少ないが、その一つひとつが特殊化していて捕獲は難しい。"),

    dict(map_id="map_atlas_mountains_005",
         name_en="Atlas Mountains",      name_ja="アトラス山脈",
         region="North Africa", biome_id="biome_mountain",
         availability_phase="initial_africa",
         unlock_rule="zoo_value", unlock_value=1500,
         recommended_hunter_rank=3, minimum_hunter_rank_gate=0,
         difficulty=3, expedition_minutes=120, base_cost_g=360, risk_level=3,
         description_en="High-altitude escarpments and Saharan foothills. Cold nights, thin trails.",
         description_ja="高地の急峻な断崖とサハラの麓。冷え込む夜、細い道。"),

    dict(map_id="map_ethiopian_highlands_006",
         name_en="Ethiopian Highlands",  name_ja="エチオピア高原",
         region="East Africa", biome_id="biome_mountain",
         availability_phase="initial_africa",
         unlock_rule="zoo_value", unlock_value=3000,
         recommended_hunter_rank=3, minimum_hunter_rank_gate=0,
         difficulty=4, expedition_minutes=180, base_cost_g=500, risk_level=3,
         description_en="Afromontane moorland and cliff. Very sparse fauna; every trip returns with something specific.",
         description_ja="アフロモンタンの荒野と断崖。動物相は極めて疎らで、遠征は具体的な狙いを持って行う。"),

    dict(map_id="map_virunga_highlands_007",
         name_en="Virunga Highlands",    name_ja="ヴィルンガ高地",
         region="Central Africa", biome_id="biome_mountain",
         availability_phase="initial_africa",
         unlock_rule="zoo_value", unlock_value=6000,
         recommended_hunter_rank=4, minimum_hunter_rank_gate=0,
         difficulty=4, expedition_minutes=300, base_cost_g=800, risk_level=4,
         description_en="Volcanic cloud forest straddling the DRC / Rwanda / Uganda border. Home of the Mountain Gorilla.",
         description_ja="コンゴ・ルワンダ・ウガンダの三国国境にまたがる火山性雲霧林。マウンテンゴリラの故郷。"),

    dict(map_id="map_kilimanjaro_slopes_008",
         name_en="Kilimanjaro Slopes",   name_ja="キリマンジャロ斜面",
         region="East Africa", biome_id="biome_mountain",
         availability_phase="initial_africa",
         unlock_rule="zoo_value", unlock_value=2000,
         recommended_hunter_rank=2, minimum_hunter_rank_gate=0,
         difficulty=3, expedition_minutes=240, base_cost_g=600, risk_level=3,
         description_en="Alpine climb from lower forest to snow line. Multiple micro-habitats; longer return.",
         description_ja="低地の森林から雪線までを登る。多様な小生息域を跨ぐが往復時間は長い。"),

    dict(map_id="map_congo_rainforest_009",
         name_en="Congo Rainforest",     name_ja="コンゴ熱帯雨林",
         region="Central Africa", biome_id="biome_rainforest",
         availability_phase="initial_africa",
         unlock_rule="zoo_value", unlock_value=5000,
         recommended_hunter_rank=4, minimum_hunter_rank_gate=0,
         difficulty=4, expedition_minutes=360, base_cost_g=900, risk_level=4,
         description_en="Dense equatorial forest. Apes, forest antelopes, and one legendary forest giraffe.",
         description_ja="赤道直下の密林。類人猿、森の偶蹄類、そして幻の森のキリンが棲む。"),

    # -- Future expansion (6) — non-Africa, not part of v0.2 playable set ----
    dict(map_id="map_amazon_rainforest_010",
         name_en="Amazon Rainforest",    name_ja="アマゾン熱帯雨林",
         region="South America", biome_id="biome_rainforest",
         availability_phase="future_expansion",
         unlock_rule="future_expansion", unlock_value=0,
         recommended_hunter_rank=3, minimum_hunter_rank_gate=0,
         difficulty=4, expedition_minutes=420, base_cost_g=1100, risk_level=4,
         description_en="Not in v0.2 initial roster. Ships when the South America expansion unlocks.",
         description_ja="v0.2の初期公開範囲には含まれない。南米エクスパンション公開時に開放。"),

    dict(map_id="map_sumatran_rainforest_011",
         name_en="Sumatran Rainforest",  name_ja="スマトラ熱帯雨林",
         region="Southeast Asia", biome_id="biome_rainforest",
         availability_phase="future_expansion",
         unlock_rule="future_expansion", unlock_value=0,
         recommended_hunter_rank=4, minimum_hunter_rank_gate=0,
         difficulty=4, expedition_minutes=480, base_cost_g=1200, risk_level=4,
         description_en="Not in v0.2. Ships with the Southeast Asia expansion (Sumatran Tiger, Sumatran Orangutan).",
         description_ja="v0.2非対応。東南アジア・エクスパンションで開放(スマトラトラ、スマトラオランウータン)。"),

    dict(map_id="map_annamite_range_012",
         name_en="Annamite Range",       name_ja="アンナン山脈",
         region="Southeast Asia", biome_id="biome_rainforest",
         availability_phase="future_expansion",
         unlock_rule="future_expansion", unlock_value=0,
         recommended_hunter_rank=6, minimum_hunter_rank_gate=0,
         difficulty=5, expedition_minutes=720, base_cost_g=2000, risk_level=5,
         description_en="Not in v0.2. Reserved for the Saola — the 'Asian Unicorn' — when SE Asia unlocks.",
         description_ja="v0.2非対応。東南アジア公開時、『アジアの一角獣』サオラのために予約されたマップ。"),

    dict(map_id="map_australian_outback_013",
         name_en="Australian Outback",   name_ja="オーストラリア奥地",
         region="Oceania", biome_id="biome_desert",
         availability_phase="future_expansion",
         unlock_rule="future_expansion", unlock_value=0,
         recommended_hunter_rank=2, minimum_hunter_rank_gate=0,
         difficulty=3, expedition_minutes=180, base_cost_g=480, risk_level=3,
         description_en="Not in v0.2. Ships with the Oceania expansion (Red Kangaroo, Dingo, Emu).",
         description_ja="v0.2非対応。オセアニア・エクスパンションで開放(アカカンガルー、ディンゴ、エミュー)。"),

    dict(map_id="map_himalayan_foothills_014",
         name_en="Himalayan Foothills",  name_ja="ヒマラヤ山麓",
         region="South Asia", biome_id="biome_mountain",
         availability_phase="future_expansion",
         unlock_rule="future_expansion", unlock_value=0,
         recommended_hunter_rank=4, minimum_hunter_rank_gate=0,
         difficulty=5, expedition_minutes=600, base_cost_g=1800, risk_level=4,
         description_en="Not in v0.2. Ships with the South Asia expansion (Snow Leopard).",
         description_ja="v0.2非対応。南アジア・エクスパンションで開放(ユキヒョウ)。"),

    dict(map_id="map_arctic_tundra_015",
         name_en="Arctic Tundra",        name_ja="北極ツンドラ",
         region="Arctic", biome_id="biome_tundra",
         availability_phase="future_expansion",
         unlock_rule="future_expansion", unlock_value=0,
         recommended_hunter_rank=4, minimum_hunter_rank_gate=0,
         difficulty=5, expedition_minutes=1200, base_cost_g=2400, risk_level=5,
         description_en="Not in v0.2. Ships with the Arctic expansion (Polar Bear, Arctic Fox).",
         description_ja="v0.2非対応。北極エクスパンションで開放(ホッキョクグマ、ホッキョクギツネ)。"),
]


# =============================================================================
# 4. Animals  (50 species; same list as v0.1 with corrections)
#
# New in v0.2:
#   availability_phase   initial_africa | future_region | special_event
#   habitat_biome_id     FK → Biomes (replaces v0.1 free-text `habitat`)
#   placement_note       explains, for future_region species, what Map is
#                        eventually required
#
# Removed in v0.2:
#   habitat  (replaced by habitat_biome_id)
# =============================================================================

# NB: category, base_zoo_value, capture_difficulty, growth_rate,
# visitor_appeal, size, active_time, description_en, description_ja
# are inherited from v0.1 unchanged unless flagged in a Review note.

ANIMALS: List[Dict[str, Any]] = [
    # ---- Common (14) ----------------------------------------------------
    dict(animal_id="animal_impala_001",             name_en="Impala",              name_ja="インパラ",
         category="hoofed",     rarity_id="rarity_common",   base_zoo_value=10,  capture_difficulty=1, growth_rate=3, visitor_appeal=8,
         habitat_biome_id="biome_savanna", size="medium", active_time="diurnal",
         availability_phase="initial_africa", placement_note="",
         description_en="Graceful savanna antelope. Everywhere on the plains.",
         description_ja="優雅なサバンナのアンテロープ。平原のどこにでもいる。"),
    dict(animal_id="animal_common_zebra_002",       name_en="Common Zebra",        name_ja="サバンナシマウマ",
         category="hoofed",     rarity_id="rarity_common",   base_zoo_value=12,  capture_difficulty=1, growth_rate=3, visitor_appeal=15,
         habitat_biome_id="biome_savanna", size="large", active_time="diurnal",
         availability_phase="initial_africa", placement_note="",
         description_en="Iconic striped equid of the East African plains.",
         description_ja="東アフリカ平原を象徴する縞模様の草食動物。"),
    dict(animal_id="animal_blue_wildebeest_003",    name_en="Blue Wildebeest",     name_ja="オグロヌー",
         category="hoofed",     rarity_id="rarity_common",   base_zoo_value=11,  capture_difficulty=1, growth_rate=3, visitor_appeal=10,
         habitat_biome_id="biome_savanna", size="large", active_time="diurnal",
         availability_phase="initial_africa", placement_note="",
         description_en="The great migrator. Millions cross the Mara every year.",
         description_ja="大移動の主役。年に数百万頭がマラ川を渡る。"),
    dict(animal_id="animal_warthog_004",            name_en="Warthog",             name_ja="イボイノシシ",
         category="hoofed",     rarity_id="rarity_common",   base_zoo_value=10,  capture_difficulty=1, growth_rate=3, visitor_appeal=12,
         habitat_biome_id="biome_savanna", size="medium", active_time="diurnal",
         availability_phase="initial_africa", placement_note="",
         description_en="Tusked pig of the savanna; runs with its tail up.",
         description_ja="サバンナに棲む牙のあるイノシシ。尾をピンと立てて走る。"),
    dict(animal_id="animal_cape_buffalo_005",       name_en="Cape Buffalo",        name_ja="アフリカスイギュウ",
         category="hoofed",     rarity_id="rarity_common",   base_zoo_value=15,  capture_difficulty=2, growth_rate=3, visitor_appeal=14,
         habitat_biome_id="biome_savanna", size="large", active_time="diurnal",
         availability_phase="initial_africa", placement_note="",
         description_en="Heavy, unpredictable herd animal. Feared even by lions.",
         description_ja="重量級で予測不能な群れ動物。ライオンさえ恐れる。"),
    dict(animal_id="animal_chacma_baboon_006",      name_en="Chacma Baboon",       name_ja="チャクマヒヒ",
         category="primate",    rarity_id="rarity_common",   base_zoo_value=10,  capture_difficulty=1, growth_rate=3, visitor_appeal=13,
         habitat_biome_id="biome_savanna", size="medium", active_time="diurnal",
         availability_phase="initial_africa", placement_note="",
         description_en="Large ground-dwelling monkey. Highly social.",
         description_ja="大型の地上性ザル。強く群れて社会を成す。"),
    dict(animal_id="animal_vervet_monkey_007",      name_en="Vervet Monkey",       name_ja="ベルベットモンキー",
         category="primate",    rarity_id="rarity_common",   base_zoo_value=8,   capture_difficulty=1, growth_rate=3, visitor_appeal=14,
         habitat_biome_id="biome_savanna", size="small", active_time="diurnal",
         availability_phase="initial_africa", placement_note="",
         description_en="Small, quick, and inquisitive.",
         description_ja="小柄で敏捷、好奇心旺盛。"),
    dict(animal_id="animal_springbok_008",          name_en="Springbok",           name_ja="スプリングボック",
         category="hoofed",     rarity_id="rarity_common",   base_zoo_value=10,  capture_difficulty=1, growth_rate=3, visitor_appeal=10,
         habitat_biome_id="biome_desert",  size="medium", active_time="diurnal",
         availability_phase="initial_africa", placement_note="",
         description_en="South African antelope famous for its vertical pronking.",
         description_ja="真上に飛び跳ねる姿で有名な南アフリカのアンテロープ。"),
    dict(animal_id="animal_meerkat_009",            name_en="Meerkat",             name_ja="ミーアキャット",
         category="small-mammal", rarity_id="rarity_common", base_zoo_value=10,  capture_difficulty=1, growth_rate=3, visitor_appeal=25,
         habitat_biome_id="biome_desert",  size="small", active_time="diurnal",
         availability_phase="initial_africa", placement_note="",
         description_en="Sentry-standing desert mongoose. Popular with visitors.",
         description_ja="立哨する砂漠のマングース。来園者に人気。"),
    dict(animal_id="animal_aardvark_010",           name_en="Aardvark",            name_ja="ツチブタ",
         category="small-mammal", rarity_id="rarity_common", base_zoo_value=12,  capture_difficulty=2, growth_rate=2, visitor_appeal=18,
         habitat_biome_id="biome_savanna", size="medium", active_time="nocturnal",
         availability_phase="initial_africa", placement_note="",
         description_en="Termite-eating burrower. Rarely seen despite being widespread.",
         description_ja="シロアリを食べる穴掘り屋。数は多いが姿を見せない。"),
    dict(animal_id="animal_hippopotamus_011",       name_en="Common Hippopotamus", name_ja="カバ",
         category="hoofed",     rarity_id="rarity_common",   base_zoo_value=14,  capture_difficulty=2, growth_rate=3, visitor_appeal=20,
         habitat_biome_id="biome_wetland", size="huge", active_time="cathemeral",
         availability_phase="initial_africa", placement_note="",
         description_en="Massive semi-aquatic. One of Africa's deadliest animals.",
         description_ja="半水生の巨体。アフリカで最も危険な動物のひとつ。"),
    dict(animal_id="animal_nile_crocodile_012",     name_en="Nile Crocodile",      name_ja="ナイルワニ",
         category="reptile",    rarity_id="rarity_common",   base_zoo_value=13,  capture_difficulty=2, growth_rate=2, visitor_appeal=15,
         habitat_biome_id="biome_wetland", size="large", active_time="diurnal",
         availability_phase="initial_africa", placement_note="",
         description_en="Ambush predator of every African river.",
         description_ja="アフリカの河川すべてに潜む待ち伏せ捕食者。"),
    dict(animal_id="animal_fennec_fox_013",         name_en="Fennec Fox",          name_ja="フェネックギツネ",
         category="small-mammal", rarity_id="rarity_common", base_zoo_value=10,  capture_difficulty=1, growth_rate=3, visitor_appeal=30,
         habitat_biome_id="biome_desert",  size="tiny", active_time="nocturnal",
         availability_phase="initial_africa", placement_note="",
         description_en="Tiny desert fox with huge ears. Sahara and Sahel margin.",
         description_ja="巨大な耳を持つ小さな砂漠のキツネ。サハラとサヘルの縁に生息。"),
    dict(animal_id="animal_dingo_014",              name_en="Dingo",               name_ja="ディンゴ",
         category="carnivore",  rarity_id="rarity_common",   base_zoo_value=11,  capture_difficulty=2, growth_rate=3, visitor_appeal=14,
         habitat_biome_id="biome_desert",  size="medium", active_time="cathemeral",
         availability_phase="future_region",
         placement_note="Requires Australian Outback map (future_expansion Oceania).",
         description_en="Australia's wild dog. Long-legged and wary.",
         description_ja="オーストラリアの野犬。脚が長く用心深い。"),

    # ---- Uncommon (14) --------------------------------------------------
    dict(animal_id="animal_masai_giraffe_015",      name_en="Masai Giraffe",       name_ja="マサイキリン",
         category="hoofed",     rarity_id="rarity_uncommon", base_zoo_value=22,  capture_difficulty=2, growth_rate=2, visitor_appeal=40,
         habitat_biome_id="biome_savanna", size="huge", active_time="diurnal",
         availability_phase="initial_africa", placement_note="",
         description_en="Tallest land animal. Every visitor wants a photo.",
         description_ja="陸上最高身長の動物。来園者は必ず撮影する。"),
    dict(animal_id="animal_african_elephant_016",   name_en="African Elephant",    name_ja="アフリカゾウ",
         category="hoofed",     rarity_id="rarity_uncommon", base_zoo_value=28,  capture_difficulty=3, growth_rate=1, visitor_appeal=50,
         habitat_biome_id="biome_savanna", size="huge", active_time="cathemeral",
         availability_phase="initial_africa", placement_note="",
         description_en="Largest land animal. Slow to age up but hugely valuable.",
         description_ja="陸上最大の動物。成長は遅いが極めて価値が高い。"),
    dict(animal_id="animal_leopard_017",            name_en="Leopard",             name_ja="ヒョウ",
         category="big-cat",    rarity_id="rarity_uncommon", base_zoo_value=25,  capture_difficulty=3, growth_rate=3, visitor_appeal=35,
         habitat_biome_id="biome_savanna", size="medium", active_time="nocturnal",
         availability_phase="initial_africa", placement_note="",
         description_en="Solitary big cat that adapts to nearly every African biome.",
         description_ja="アフリカのほぼ全バイオームに適応する単独行動の大型ネコ科。"),
    dict(animal_id="animal_cheetah_018",            name_en="Cheetah",             name_ja="チーター",
         category="big-cat",    rarity_id="rarity_uncommon", base_zoo_value=24,  capture_difficulty=3, growth_rate=3, visitor_appeal=38,
         habitat_biome_id="biome_savanna", size="medium", active_time="diurnal",
         availability_phase="initial_africa", placement_note="",
         description_en="Fastest land animal. Fragile compared to other big cats.",
         description_ja="陸上最速の動物。他の大型ネコ科より繊細。"),
    dict(animal_id="animal_spotted_hyena_019",      name_en="Spotted Hyena",       name_ja="ブチハイエナ",
         category="carnivore",  rarity_id="rarity_uncommon", base_zoo_value=20,  capture_difficulty=2, growth_rate=3, visitor_appeal=18,
         habitat_biome_id="biome_savanna", size="medium", active_time="nocturnal",
         availability_phase="initial_africa", placement_note="",
         description_en="Highly social carnivore. Clan-based societies.",
         description_ja="高度に社会的な捕食者。氏族制の群れをつくる。"),
    dict(animal_id="animal_african_wild_dog_020",   name_en="African Wild Dog",    name_ja="リカオン",
         category="carnivore",  rarity_id="rarity_uncommon", base_zoo_value=22,  capture_difficulty=3, growth_rate=3, visitor_appeal=22,
         habitat_biome_id="biome_savanna", size="medium", active_time="diurnal",
         availability_phase="initial_africa", placement_note="",
         description_en="Endangered pack hunter with the highest hunting success rate on the continent.",
         description_ja="絶滅危惧。大陸屈指の狩猟成功率を誇る群れハンター。"),
    dict(animal_id="animal_serval_021",             name_en="Serval",              name_ja="サーバル",
         category="small-mammal", rarity_id="rarity_uncommon", base_zoo_value=18, capture_difficulty=2, growth_rate=3, visitor_appeal=28,
         habitat_biome_id="biome_savanna", size="small", active_time="nocturnal",
         availability_phase="initial_africa", placement_note="",
         description_en="Long-legged small cat. Leaps to catch birds.",
         description_ja="脚の長い小型ネコ科。跳躍して鳥を捕える。"),
    dict(animal_id="animal_caracal_022",            name_en="Caracal",             name_ja="カラカル",
         category="small-mammal", rarity_id="rarity_uncommon", base_zoo_value=18, capture_difficulty=2, growth_rate=3, visitor_appeal=26,
         habitat_biome_id="biome_mountain", size="small", active_time="nocturnal",
         availability_phase="initial_africa", placement_note="",
         description_en="Tufted-eared cat of arid and mountainous zones — real range includes North Africa.",
         description_ja="房耳のネコ科。乾燥地帯と山岳。実生息域は北アフリカを含む。"),
    dict(animal_id="animal_red_kangaroo_023",       name_en="Red Kangaroo",        name_ja="アカカンガルー",
         category="marsupial",  rarity_id="rarity_uncommon", base_zoo_value=20,  capture_difficulty=2, growth_rate=3, visitor_appeal=32,
         habitat_biome_id="biome_desert",  size="large", active_time="crepuscular",
         availability_phase="future_region",
         placement_note="Requires Australian Outback map (future_expansion Oceania).",
         description_en="World's largest marsupial. Australia only.",
         description_ja="世界最大の有袋類。オーストラリアのみ。"),
    dict(animal_id="animal_ethiopian_wolf_024",     name_en="Ethiopian Wolf",      name_ja="エチオピアオオカミ",
         category="carnivore",  rarity_id="rarity_uncommon", base_zoo_value=22,  capture_difficulty=3, growth_rate=3, visitor_appeal=25,
         habitat_biome_id="biome_mountain", size="medium", active_time="diurnal",
         availability_phase="initial_africa", placement_note="",
         description_en="Endangered highland canid endemic to Ethiopia. Only ~500 left.",
         description_ja="エチオピア高原固有の絶滅危惧のイヌ科。生息数は約500頭。"),
    dict(animal_id="animal_barbary_macaque_025",    name_en="Barbary Macaque",     name_ja="バーバリーマカク",
         category="primate",    rarity_id="rarity_uncommon", base_zoo_value=18,  capture_difficulty=2, growth_rate=3, visitor_appeal=22,
         habitat_biome_id="biome_mountain", size="medium", active_time="diurnal",
         availability_phase="initial_africa", placement_note="",
         description_en="Only macaque native to Africa. Atlas cedar forests.",
         description_ja="アフリカ産唯一のマカク。アトラス山脈のシダー林に棲む。"),
    dict(animal_id="animal_emu_026",                name_en="Emu",                 name_ja="エミュー",
         category="bird",       rarity_id="rarity_uncommon", base_zoo_value=18,  capture_difficulty=2, growth_rate=3, visitor_appeal=24,
         habitat_biome_id="biome_desert",  size="large", active_time="diurnal",
         availability_phase="future_region",
         placement_note="Requires Australian Outback map (future_expansion Oceania).",
         description_en="Second-tallest living bird. Curious and fast.",
         description_ja="現生では二番目に背の高い鳥。好奇心旺盛で走るのが速い。"),
    dict(animal_id="animal_bat_eared_fox_027",      name_en="Bat-eared Fox",       name_ja="オオミミギツネ",
         category="small-mammal", rarity_id="rarity_uncommon", base_zoo_value=16, capture_difficulty=2, growth_rate=3, visitor_appeal=26,
         habitat_biome_id="biome_savanna", size="small", active_time="cathemeral",
         availability_phase="initial_africa", placement_note="",
         description_en="Insect-eating fox with cartoon-large ears.",
         description_ja="漫画のような大きな耳を持つ、昆虫を食べるキツネ。"),
    dict(animal_id="animal_marabou_stork_028",      name_en="Marabou Stork",       name_ja="アフリカハゲコウ",
         category="bird",       rarity_id="rarity_uncommon", base_zoo_value=17,  capture_difficulty=2, growth_rate=3, visitor_appeal=15,
         habitat_biome_id="biome_wetland", size="large", active_time="diurnal",
         availability_phase="initial_africa", placement_note="",
         description_en="Massive African scavenger. 3.5-metre wingspan.",
         description_ja="巨大なアフリカのスカベンジャー。翼開長3.5m。"),

    # ---- Rare (10) ------------------------------------------------------
    dict(animal_id="animal_lion_029",               name_en="African Lion",        name_ja="ライオン",
         category="big-cat",    rarity_id="rarity_rare",     base_zoo_value=45,  capture_difficulty=4, growth_rate=2, visitor_appeal=70,
         habitat_biome_id="biome_savanna", size="large", active_time="nocturnal",
         availability_phase="initial_africa", placement_note="",
         description_en="Only truly social big cat. Icon of the savanna.",
         description_ja="真に社会性を持つ唯一の大型ネコ科。サバンナの象徴。"),
    dict(animal_id="animal_black_rhinoceros_030",   name_en="Black Rhinoceros",    name_ja="クロサイ",
         category="hoofed",     rarity_id="rarity_rare",     base_zoo_value=50,  capture_difficulty=4, growth_rate=1, visitor_appeal=55,
         habitat_biome_id="biome_savanna", size="huge", active_time="nocturnal",
         availability_phase="initial_africa", placement_note="",
         description_en="Critically endangered. Pointed prehensile lip for browsing.",
         description_ja="絶滅寸前。伸縮する尖った上唇で葉を食む。"),
    dict(animal_id="animal_chimpanzee_031",         name_en="Chimpanzee",          name_ja="チンパンジー",
         category="primate",    rarity_id="rarity_rare",     base_zoo_value=42,  capture_difficulty=4, growth_rate=2, visitor_appeal=65,
         habitat_biome_id="biome_rainforest", size="medium", active_time="diurnal",
         availability_phase="initial_africa", placement_note="",
         description_en="Closest living relative to humans. Tool-users.",
         description_ja="ヒトに最も近い現生種のひとつ。道具を使う。"),
    dict(animal_id="animal_mountain_gorilla_032",   name_en="Mountain Gorilla",    name_ja="マウンテンゴリラ",
         category="primate",    rarity_id="rarity_rare",     base_zoo_value=55,  capture_difficulty=4, growth_rate=1, visitor_appeal=75,
         habitat_biome_id="biome_mountain", size="huge", active_time="diurnal",
         availability_phase="initial_africa", placement_note="",
         description_en="Virunga cloud-forest gorilla. Around 1,000 left in the wild.",
         description_ja="ヴィルンガ雲霧林のゴリラ。野生では約1,000頭のみ。"),
    dict(animal_id="animal_bongo_033",              name_en="Bongo",               name_ja="ボンゴ",
         category="hoofed",     rarity_id="rarity_rare",     base_zoo_value=38,  capture_difficulty=4, growth_rate=2, visitor_appeal=45,
         habitat_biome_id="biome_rainforest", size="large", active_time="nocturnal",
         availability_phase="initial_africa", placement_note="",
         description_en="Striped forest antelope of the Congo basin. Almost never seen in the wild.",
         description_ja="コンゴ盆地の縞模様の森林アンテロープ。野生でその姿はほぼ見られない。"),
    dict(animal_id="animal_jaguar_034",             name_en="Jaguar",              name_ja="ジャガー",
         category="big-cat",    rarity_id="rarity_rare",     base_zoo_value=48,  capture_difficulty=4, growth_rate=2, visitor_appeal=60,
         habitat_biome_id="biome_rainforest", size="large", active_time="cathemeral",
         availability_phase="future_region",
         placement_note="Requires Amazon Rainforest map (future_expansion South America).",
         description_en="Third-largest big cat and Amazon apex predator.",
         description_ja="3番目に大きな大型ネコ科。アマゾンの頂点捕食者。"),
    dict(animal_id="animal_giant_anteater_035",     name_en="Giant Anteater",      name_ja="オオアリクイ",
         category="small-mammal", rarity_id="rarity_rare",   base_zoo_value=35,  capture_difficulty=3, growth_rate=2, visitor_appeal=40,
         habitat_biome_id="biome_rainforest", size="large", active_time="diurnal",
         availability_phase="future_region",
         placement_note="Requires Amazon Rainforest map (future_expansion South America).",
         description_en="2-metre insect specialist. Distinctive coat pattern.",
         description_ja="全長2mの昆虫食スペシャリスト。特徴的な体色。"),
    dict(animal_id="animal_snow_leopard_036",       name_en="Snow Leopard",        name_ja="ユキヒョウ",
         category="big-cat",    rarity_id="rarity_rare",     base_zoo_value=55,  capture_difficulty=5, growth_rate=2, visitor_appeal=75,
         habitat_biome_id="biome_mountain", size="medium", active_time="crepuscular",
         availability_phase="future_region",
         placement_note="Requires Himalayan Foothills map (future_expansion South Asia).",
         description_en="Ghost of the mountains. Legendary elusiveness in real life.",
         description_ja="山の亡霊。実生活でも伝説的なほど姿を見せない。"),
    dict(animal_id="animal_giant_panda_037",        name_en="Giant Panda",         name_ja="ジャイアントパンダ",
         category="carnivore",  rarity_id="rarity_rare",     base_zoo_value=60,  capture_difficulty=4, growth_rate=1, visitor_appeal=90,
         habitat_biome_id="biome_mountain", size="large", active_time="cathemeral",
         availability_phase="future_region",
         placement_note="No matching map in v0.2. Real range = Sichuan bamboo forests. Requires a future East Asia map.",
         description_en="Bamboo-eating bear. Highest visitor draw at Rare tier.",
         description_ja="竹を食べるクマ科。レア帯で最大級の集客力。"),
    dict(animal_id="animal_bengal_tiger_038",       name_en="Bengal Tiger",        name_ja="ベンガルトラ",
         category="big-cat",    rarity_id="rarity_rare",     base_zoo_value=52,  capture_difficulty=4, growth_rate=2, visitor_appeal=72,
         habitat_biome_id="biome_rainforest", size="large", active_time="nocturnal",
         availability_phase="future_region",
         placement_note="No matching map in v0.2. Real range = Indian subcontinent lowlands. Requires a future Indian Subcontinent map.",
         description_en="Largest wild cat. Solitary and powerful.",
         description_ja="現生最大の野生ネコ科。単独行動で強靭。"),

    # ---- Epic (7) -------------------------------------------------------
    dict(animal_id="animal_northern_white_rhino_039", name_en="Northern White Rhinoceros", name_ja="キタシロサイ",
         category="hoofed",     rarity_id="rarity_epic",     base_zoo_value=110, capture_difficulty=5, growth_rate=1, visitor_appeal=95,
         habitat_biome_id="biome_savanna", size="huge", active_time="diurnal",
         availability_phase="special_event",
         placement_note="Removed from normal spawn (v0.2 review). Reserved for future Conservation / Special Event surface.",
         description_en="Functionally extinct in the wild (2 individuals remain). Reserved for a future event.",
         description_ja="野生では事実上絶滅(現存2頭)。将来イベント用に予約。"),
    dict(animal_id="animal_bonobo_040",             name_en="Bonobo",              name_ja="ボノボ",
         category="primate",    rarity_id="rarity_epic",     base_zoo_value=85,  capture_difficulty=5, growth_rate=2, visitor_appeal=80,
         habitat_biome_id="biome_rainforest", size="medium", active_time="diurnal",
         availability_phase="initial_africa", placement_note="",
         description_en="Peaceful great ape. Only south of the Congo River.",
         description_ja="穏やかな大型類人猿。コンゴ川の南側にのみ生息。"),
    dict(animal_id="animal_sumatran_tiger_041",     name_en="Sumatran Tiger",      name_ja="スマトラトラ",
         category="big-cat",    rarity_id="rarity_epic",     base_zoo_value=95,  capture_difficulty=5, growth_rate=2, visitor_appeal=88,
         habitat_biome_id="biome_rainforest", size="medium", active_time="nocturnal",
         availability_phase="future_region",
         placement_note="Requires Sumatran Rainforest map (future_expansion Southeast Asia).",
         description_en="Smallest tiger subspecies, less than 400 left.",
         description_ja="最も小型のトラ亜種。400頭未満しか残っていない。"),
    dict(animal_id="animal_sumatran_orangutan_042", name_en="Sumatran Orangutan",  name_ja="スマトラオランウータン",
         category="primate",    rarity_id="rarity_epic",     base_zoo_value=90,  capture_difficulty=4, growth_rate=1, visitor_appeal=82,
         habitat_biome_id="biome_rainforest", size="large", active_time="diurnal",
         availability_phase="future_region",
         placement_note="Requires Sumatran Rainforest map (future_expansion Southeast Asia).",
         description_en="Red ape of the canopy. Critically endangered.",
         description_ja="樹冠に棲む赤い類人猿。近絶滅種。"),
    dict(animal_id="animal_polar_bear_043",         name_en="Polar Bear",          name_ja="ホッキョクグマ",
         category="carnivore",  rarity_id="rarity_epic",     base_zoo_value=95,  capture_difficulty=5, growth_rate=2, visitor_appeal=85,
         habitat_biome_id="biome_tundra",  size="huge", active_time="cathemeral",
         availability_phase="future_region",
         placement_note="Requires Arctic Tundra map (future_expansion Arctic).",
         description_en="Ice-dependent apex predator.",
         description_ja="氷に依存する頂点捕食者。"),
    dict(animal_id="animal_arctic_fox_044",         name_en="Arctic Fox",          name_ja="ホッキョクギツネ",
         category="small-mammal", rarity_id="rarity_epic",   base_zoo_value=75,  capture_difficulty=4, growth_rate=3, visitor_appeal=70,
         habitat_biome_id="biome_tundra",  size="small", active_time="cathemeral",
         availability_phase="future_region",
         placement_note="Requires Arctic Tundra map (future_expansion Arctic).",
         description_en="White in winter, brown in summer. Champion of insulation.",
         description_ja="冬は白、夏は茶。断熱性能の王者。"),
    dict(animal_id="animal_amur_leopard_045",       name_en="Amur Leopard",        name_ja="アムールヒョウ",
         category="big-cat",    rarity_id="rarity_epic",     base_zoo_value=95,  capture_difficulty=5, growth_rate=2, visitor_appeal=85,
         habitat_biome_id="biome_mountain", size="medium", active_time="nocturnal",
         availability_phase="future_region",
         placement_note="No matching map in v0.2. Real range = Russian Far East / NE China. Requires a future Amur / Manchurian map.",
         description_en="Cold-adapted leopard. Fewer than 120 individuals globally.",
         description_ja="寒冷地適応のヒョウ亜種。世界に120頭未満。"),

    # ---- Legendary (5) --------------------------------------------------
    dict(animal_id="animal_saola_046",              name_en="Saola",               name_ja="サオラ",
         category="hoofed",     rarity_id="rarity_legendary", base_zoo_value=220, capture_difficulty=5, growth_rate=1, visitor_appeal=90,
         habitat_biome_id="biome_rainforest", size="medium", active_time="nocturnal",
         availability_phase="future_region",
         placement_note="Requires Annamite Range map (future_expansion Southeast Asia).",
         description_en="The 'Asian Unicorn'. Almost mythical Vietnamese bovid.",
         description_ja="『アジアの一角獣』。ほぼ神話的なベトナムのウシ科動物。"),
    dict(animal_id="animal_javan_rhinoceros_047",   name_en="Javan Rhinoceros",    name_ja="ジャワサイ",
         category="hoofed",     rarity_id="rarity_legendary", base_zoo_value=260, capture_difficulty=5, growth_rate=1, visitor_appeal=95,
         habitat_biome_id="biome_rainforest", size="huge", active_time="cathemeral",
         availability_phase="future_region",
         placement_note="No matching map in v0.2. Real range = Ujung Kulon NP, Java. Requires a future Javan Rainforest map.",
         description_en="Around 76 individuals total.",
         description_ja="現存総数約76頭。"),
    dict(animal_id="animal_kakapo_048",             name_en="Kakapo",              name_ja="カカポ",
         category="bird",       rarity_id="rarity_legendary", base_zoo_value=200, capture_difficulty=4, growth_rate=1, visitor_appeal=88,
         habitat_biome_id="biome_rainforest", size="medium", active_time="nocturnal",
         availability_phase="future_region",
         placement_note="No matching map in v0.2. Real range = predator-free islands off New Zealand. Requires a future New Zealand map.",
         description_en="Flightless nocturnal parrot.",
         description_ja="飛べない夜行性のオウム。"),
    dict(animal_id="animal_okapi_049",              name_en="Okapi",               name_ja="オカピ",
         category="hoofed",     rarity_id="rarity_legendary", base_zoo_value=210, capture_difficulty=5, growth_rate=1, visitor_appeal=92,
         habitat_biome_id="biome_rainforest", size="large", active_time="diurnal",
         availability_phase="initial_africa", placement_note="",
         description_en="Congo's forest giraffe. Ituri Forest, DRC.",
         description_ja="コンゴの森のキリン。DRC・イトゥリ森林。"),
    dict(animal_id="animal_cross_river_gorilla_050", name_en="Cross River Gorilla", name_ja="クロスリバーゴリラ",
         category="primate",    rarity_id="rarity_legendary", base_zoo_value=240, capture_difficulty=5, growth_rate=1, visitor_appeal=94,
         habitat_biome_id="biome_rainforest", size="huge", active_time="diurnal",
         availability_phase="future_region",
         placement_note="No matching map in v0.2. Real range = Cameroon-Nigeria border. Requires a future West African Rainforest map.",
         description_en="Rarest great ape. Fewer than 300 individuals.",
         description_ja="最も希少な類人猿。300頭未満。"),
]

ANIMAL_INDEX = {a["animal_id"]: a for a in ANIMALS}


# =============================================================================
# 5. Hunters  (18; contract_cost_g rename; 3 fictionalised names)
#
# Rank progression (Draft, retained from v0.1 — see Review for status):
#   Bronze → Silver → Gold → Platinum → Diamond → Master
#
# capture_bonus / rare_find_bonus / speed_bonus are additive percentage
# points on top of ExpeditionRules base values.
#
# preferred_biome:  a biome_id OR the literal string "any" (generalist).
#
# Contract model: one-off Guild-pool contract per expedition. NOT owned
# by the player. NOT persistent inventory. See README §"HunterはPlayer所有しない".
# =============================================================================

HUNTERS: List[Dict[str, Any]] = [
    # -- Beginner tier (3) --
    dict(hunter_id="hunter_amara_kone_001",   name="Amara Koné",
         rank="Bronze", level=1, specialty="Beginner",
         preferred_biome_id="biome_savanna", capture_bonus=0,  rare_find_bonus=-5, speed_bonus=0,
         contract_cost_g=50,
         personality="Cheerful new recruit from the coastal towns.",
         description="Reliable first Hunter for any player. Cheap, competent, no surprises."),

    dict(hunter_id="hunter_kofi_mensah_002",  name="Kofi Mensah",
         rank="Bronze", level=2, specialty="Beginner-Forest",
         preferred_biome_id="biome_rainforest", capture_bonus=0, rare_find_bonus=-5, speed_bonus=0,
         contract_cost_g=60,
         personality="Grew up on the edge of the Congo. Quiet but sure-footed.",
         description="Second-cheapest Hunter. Slight edge in forest biomes."),

    dict(hunter_id="hunter_hana_ito_003",     name="Hana Ito",
         rank="Bronze", level=3, specialty="All-round-Rookie",
         preferred_biome_id="any", capture_bonus=+2, rare_find_bonus=0, speed_bonus=0,
         contract_cost_g=90,
         personality="Level-headed rookie from an old expedition family.",
         description="Slightly stronger than the other rookies; often the second Hunter a player unlocks."),

    # -- Rare-find specialists (2) --
    dict(hunter_id="hunter_zara_okafor_004",  name="Zara Okafor",
         rank="Silver", level=5, specialty="Rare-Find",
         preferred_biome_id="biome_savanna", capture_bonus=0, rare_find_bonus=+15, speed_bonus=-5,
         contract_cost_g=280,
         personality="Patient tracker. Would rather come back empty-handed than settle for a common find.",
         description="Boosts the odds of returning with a Rare or better, at the cost of speed."),

    dict(hunter_id="hunter_diego_ramirez_005", name="Diego Ramírez",
         rank="Silver", level=6, specialty="Rare-Find-Rainforest",
         preferred_biome_id="biome_rainforest", capture_bonus=0, rare_find_bonus=+18, speed_bonus=-5,
         contract_cost_g=340,
         personality="Amazon-born biologist. Reads the canopy like a book.",
         description="Rare-find specialist tuned for rainforest maps."),

    # -- Mountain / high-altitude specialists (2) --
    dict(hunter_id="hunter_nadia_kowalski_006", name="Nadia Kowalski",
         rank="Silver", level=5, specialty="Mountain",
         preferred_biome_id="biome_mountain", capture_bonus=+10, rare_find_bonus=0, speed_bonus=0,
         contract_cost_g=260,
         personality="Alpine specialist. Speaks softly, climbs faster than anyone.",
         description="Substantially better on mountain maps; average elsewhere."),

    # -- v0.2 fictional rename: v0.1 'Tenzing Sherpa' evoked a specific
    #    real historical mountaineer. Replaced with an invented name that
    #    still fits the high-altitude specialist archetype without
    #    mapping to any specific real person.
    dict(hunter_id="hunter_nima_kirat_007",   name="Nima Kirat",
         rank="Gold", level=7, specialty="High-Altitude",
         preferred_biome_id="biome_mountain", capture_bonus=+15, rare_find_bonus=+5, speed_bonus=-10,
         contract_cost_g=780,
         personality="Legendary high-altitude guide. The only Hunter fluent above 6000m.",
         description="Only Hunter who reliably returns from extreme mountain expeditions."),

    # -- Nocturnal specialists (2) --
    dict(hunter_id="hunter_yuki_nakamura_008", name="Yuki Nakamura",
         rank="Silver", level=6, specialty="Nocturnal",
         preferred_biome_id="any", capture_bonus=+8, rare_find_bonus=+8, speed_bonus=0,
         contract_cost_g=320,
         personality="Sleeps by day, hunts by lantern-light.",
         description="Bonuses apply to nocturnal species regardless of biome (rule TBD)."),

    # -- v0.2 fictional rename: 'Redcloud' echoed a specific historical
    #    Native American leader. Replaced with an invented family name.
    dict(hunter_id="hunter_sky_winterhawk_009", name="Sky Winterhawk",
         rank="Gold", level=7, specialty="Nocturnal-Wilderness",
         preferred_biome_id="biome_tundra", capture_bonus=+10, rare_find_bonus=+10, speed_bonus=0,
         contract_cost_g=520,
         personality="Trained in the northern boreal. Reads tracks by moonlight.",
         description="Nocturnal specialist tuned for cold biomes."),

    # -- High-difficulty capture (2) --
    # -- v0.2 fictional rename: 'Chandra' overlaps a well-known scientist
    #    surname. Replaced with a common but non-eponymous family name.
    dict(hunter_id="hunter_ravi_menon_010",   name="Ravi Menon",
         rank="Gold", level=8, specialty="Hard-Capture-Tropical",
         preferred_biome_id="biome_rainforest", capture_bonus=+20, rare_find_bonus=0, speed_bonus=0,
         contract_cost_g=680,
         personality="Ex-military tracker. Nothing runs from him twice.",
         description="Straight-up better at succeeding on hard maps; no rare-find bias."),

    dict(hunter_id="hunter_elena_marchetti_011", name="Elena Marchetti",
         rank="Platinum", level=9, specialty="Hard-Capture-Universal",
         preferred_biome_id="any", capture_bonus=+25, rare_find_bonus=+5, speed_bonus=-5,
         contract_cost_g=1400,
         personality="Tracker who works in every biome.",
         description="Elite generalist. High cost; near-guaranteed success on high-difficulty maps."),

    # -- High-speed expedition (2) --
    dict(hunter_id="hunter_miguel_santos_012", name="Miguel Santos",
         rank="Silver", level=5, specialty="Speed",
         preferred_biome_id="any", capture_bonus=-5, rare_find_bonus=-5, speed_bonus=+30,
         contract_cost_g=240,
         personality="Runs everywhere. Returns before you finish your coffee.",
         description="Best-in-slot for cycling many short expeditions per real-world day."),

    dict(hunter_id="hunter_sara_lindqvist_013", name="Sara Lindqvist",
         rank="Gold", level=7, specialty="Speed-Cold",
         preferred_biome_id="biome_tundra", capture_bonus=0, rare_find_bonus=0, speed_bonus=+25,
         contract_cost_g=560,
         personality="Skis where others walk. Loves the cold; hates the tropics.",
         description="Speed specialist tuned for tundra maps; slight penalty on hot maps (rule TBD)."),

    # -- Balanced (3) --
    dict(hunter_id="hunter_chen_wei_014",     name="Chen Wei",
         rank="Silver", level=6, specialty="Balanced",
         preferred_biome_id="any", capture_bonus=+5, rare_find_bonus=+3, speed_bonus=+3,
         contract_cost_g=300,
         personality="Steady. Reliable. Talks little.",
         description="Balanced upgrade over rookies. First mid-tier Hunter most players buy."),

    dict(hunter_id="hunter_priya_kaur_015",   name="Priya Kaur",
         rank="Gold", level=7, specialty="Balanced-Wetland",
         preferred_biome_id="biome_wetland", capture_bonus=+8, rare_find_bonus=+3, speed_bonus=+3,
         contract_cost_g=540,
         personality="Grew up on the Delta. Reads water.",
         description="Balanced Gold-tier Hunter with wetland edge."),

    dict(hunter_id="hunter_kwame_boateng_016", name="Kwame Boateng",
         rank="Gold", level=8, specialty="Balanced-Savanna",
         preferred_biome_id="biome_savanna", capture_bonus=+8, rare_find_bonus=+3, speed_bonus=+3,
         contract_cost_g=580,
         personality="Guild instructor turned field Hunter.",
         description="Balanced Gold-tier Hunter with savanna edge."),

    # -- Legendary tier (2, shared scarce) --
    # -- v0.2 fictional rename: v0.1 'Fujimori' overlapped a real political
    #    surname. Replaced with a common Japanese family name without
    #    similar overlap.
    dict(hunter_id="hunter_aiko_tanabe_017",  name="Aiko Tanabe",
         rank="Diamond", level=10, specialty="Legendary-All-Round",
         preferred_biome_id="any", capture_bonus=+20, rare_find_bonus=+15, speed_bonus=+10,
         contract_cost_g=3800,
         personality="Undefeated in three continents. Books herself.",
         description="Shared scarce Hunter. Only one player may hold her contract at a time."),

    dict(hunter_id="hunter_dr_malik_osei_018", name="Dr. Malik Osei",
         rank="Master", level=10, specialty="Legendary-Rare-Find",
         preferred_biome_id="biome_rainforest", capture_bonus=+15, rare_find_bonus=+30, speed_bonus=0,
         contract_cost_g=4500,
         personality="Field biologist. Has found the Saola. Once.",
         description="Shared scarce Hunter. Highest rare-find bonus in the game."),
]


# =============================================================================
# 6. MapAnimals  (real-habitat only)
#
# NB: v0.2 removes every game-abstraction row from v0.1
# (Kakapo→Borneo, Saola→Borneo, Bengal Tiger→Borneo, Mountain Gorilla→
#  Kilimanjaro, Ethiopian Wolf→Atlas, Giant Panda→Himalayan, Amur
#  Leopard→Himalayan, Cross River Gorilla→Congo). Those animals now live
#  in `availability_phase = future_region` and are not spawned on any
#  Map until the required expansion Map ships.
#
# Northern White Rhinoceros has NO row: it is `special_event`.
# =============================================================================

def _ma(map_id: str, animal_id: str, spawn_weight: int,
        capture_mod: int = 0, notes: str = "",
        needs_review: bool = False) -> Dict[str, Any]:
    return dict(map_id=map_id, animal_id=animal_id, spawn_weight=spawn_weight,
                capture_modifier=capture_mod, notes=notes,
                needs_review=needs_review)


MAP_ANIMALS_RAW: List[Dict[str, Any]] = [
    # -- Kenyan Savanna --
    _ma("map_kenyan_savanna_001",   "animal_impala_001",             35),
    _ma("map_kenyan_savanna_001",   "animal_common_zebra_002",       30),
    _ma("map_kenyan_savanna_001",   "animal_blue_wildebeest_003",    28),
    _ma("map_kenyan_savanna_001",   "animal_warthog_004",            30),
    _ma("map_kenyan_savanna_001",   "animal_cape_buffalo_005",       20),
    _ma("map_kenyan_savanna_001",   "animal_chacma_baboon_006",      15,
        notes="Chacma is southern; olive baboon in East Africa. Genus-level abstraction.",
        needs_review=True),
    _ma("map_kenyan_savanna_001",   "animal_vervet_monkey_007",      25),
    _ma("map_kenyan_savanna_001",   "animal_aardvark_010",            8, notes="nocturnal — bonus at night"),
    _ma("map_kenyan_savanna_001",   "animal_masai_giraffe_015",      15),
    _ma("map_kenyan_savanna_001",   "animal_african_elephant_016",   10),
    _ma("map_kenyan_savanna_001",   "animal_leopard_017",             8),
    _ma("map_kenyan_savanna_001",   "animal_cheetah_018",             8),
    _ma("map_kenyan_savanna_001",   "animal_spotted_hyena_019",      14),
    _ma("map_kenyan_savanna_001",   "animal_bat_eared_fox_027",       9),
    _ma("map_kenyan_savanna_001",   "animal_lion_029",                6),
    _ma("map_kenyan_savanna_001",   "animal_black_rhinoceros_030",    3),

    # -- Serengeti Plains --
    _ma("map_serengeti_plains_002", "animal_impala_001",             30),
    _ma("map_serengeti_plains_002", "animal_common_zebra_002",       35),
    _ma("map_serengeti_plains_002", "animal_blue_wildebeest_003",    40, notes="peak during migration season"),
    _ma("map_serengeti_plains_002", "animal_warthog_004",            25),
    _ma("map_serengeti_plains_002", "animal_cape_buffalo_005",       22),
    _ma("map_serengeti_plains_002", "animal_chacma_baboon_006",      12,
        notes="genus-level abstraction, see notes on map_kenyan_savanna_001",
        needs_review=True),
    _ma("map_serengeti_plains_002", "animal_masai_giraffe_015",      18),
    _ma("map_serengeti_plains_002", "animal_african_elephant_016",   14),
    _ma("map_serengeti_plains_002", "animal_leopard_017",            10),
    _ma("map_serengeti_plains_002", "animal_cheetah_018",            14),
    _ma("map_serengeti_plains_002", "animal_spotted_hyena_019",      18),
    _ma("map_serengeti_plains_002", "animal_african_wild_dog_020",    8),
    _ma("map_serengeti_plains_002", "animal_lion_029",                9),
    _ma("map_serengeti_plains_002", "animal_black_rhinoceros_030",    3),

    # -- Okavango Delta --
    _ma("map_okavango_delta_003",   "animal_cape_buffalo_005",       22),
    _ma("map_okavango_delta_003",   "animal_hippopotamus_011",       28),
    _ma("map_okavango_delta_003",   "animal_nile_crocodile_012",     26),
    _ma("map_okavango_delta_003",   "animal_african_elephant_016",   16),
    _ma("map_okavango_delta_003",   "animal_serval_021",             14),
    _ma("map_okavango_delta_003",   "animal_african_wild_dog_020",    9),
    _ma("map_okavango_delta_003",   "animal_marabou_stork_028",      18),

    # -- Namib Desert --
    _ma("map_namib_desert_004",     "animal_springbok_008",          28),
    _ma("map_namib_desert_004",     "animal_meerkat_009",            32),
    _ma("map_namib_desert_004",     "animal_aardvark_010",           10, notes="nocturnal"),
    _ma("map_namib_desert_004",     "animal_bat_eared_fox_027",      18),

    # -- Atlas Mountains --
    _ma("map_atlas_mountains_005",  "animal_barbary_macaque_025",    30),
    _ma("map_atlas_mountains_005",  "animal_caracal_022",            18, notes="Real range includes North Africa Atlas."),
    _ma("map_atlas_mountains_005",  "animal_fennec_fox_013",         14, notes="Saharan margin — Atlas foothills serve as a proxy edge in v0.2.",
        needs_review=True),

    # -- Ethiopian Highlands (new v0.2 map) --
    _ma("map_ethiopian_highlands_006", "animal_ethiopian_wolf_024",  40, notes="Endemic to the Ethiopian Highlands. Signature species of this map."),
    _ma("map_ethiopian_highlands_006", "animal_chacma_baboon_006",   10,
        notes="Gelada would be the real fit; roster does not include it in v0.2. Chacma placeholder pending v0.3.",
        needs_review=True),

    # -- Virunga Highlands (new v0.2 map) --
    _ma("map_virunga_highlands_007", "animal_mountain_gorilla_032",  25, notes="Signature species — Virunga Massif is Mountain Gorilla's real range."),
    _ma("map_virunga_highlands_007", "animal_chimpanzee_031",         8, notes="Eastern Chimpanzee real range overlaps the Virunga forests."),

    # -- Kilimanjaro Slopes --
    _ma("map_kilimanjaro_slopes_008", "animal_leopard_017",           8, notes="Mountain-adapted leopard population."),
    _ma("map_kilimanjaro_slopes_008", "animal_cape_buffalo_005",     10, notes="Highland herds documented on lower slopes."),
    _ma("map_kilimanjaro_slopes_008", "animal_chacma_baboon_006",    12,
        notes="genus-level abstraction — real species on Kilimanjaro is olive baboon.",
        needs_review=True),
    _ma("map_kilimanjaro_slopes_008", "animal_african_elephant_016",  8, notes="Lower montane forest elephants."),

    # -- Congo Rainforest --
    _ma("map_congo_rainforest_009", "animal_leopard_017",             8),
    _ma("map_congo_rainforest_009", "animal_chimpanzee_031",         14),
    _ma("map_congo_rainforest_009", "animal_bongo_033",              10, notes="Central African forest range."),
    _ma("map_congo_rainforest_009", "animal_bonobo_040",              6, notes="South of the Congo River only."),
    _ma("map_congo_rainforest_009", "animal_okapi_049",               2, capture_mod=-15, notes="Ituri Forest, DRC — real range."),

    # -- Future_expansion maps get their intended roster now, but only
    #    activate when the expansion ships. Rows are here so the game
    #    knows what should spawn when the Map's availability_phase flips.

    # -- Amazon Rainforest (future) --
    _ma("map_amazon_rainforest_010", "animal_jaguar_034",            12, notes="Apex predator of the Amazon basin."),
    _ma("map_amazon_rainforest_010", "animal_giant_anteater_035",    16),

    # -- Sumatran Rainforest (future) --
    _ma("map_sumatran_rainforest_011", "animal_sumatran_tiger_041",   4, notes="Sumatran endemic."),
    _ma("map_sumatran_rainforest_011", "animal_sumatran_orangutan_042", 6, notes="Sumatran endemic."),

    # -- Annamite Range (future) --
    _ma("map_annamite_range_012",     "animal_saola_046",             1, capture_mod=-30, notes="Vietnam/Laos Annamite endemic."),

    # -- Australian Outback (future) --
    _ma("map_australian_outback_013", "animal_red_kangaroo_023",     32),
    _ma("map_australian_outback_013", "animal_dingo_014",            28),
    _ma("map_australian_outback_013", "animal_emu_026",              26),

    # -- Himalayan Foothills (future) --
    _ma("map_himalayan_foothills_014", "animal_snow_leopard_036",    10, notes="High-altitude Himalayan range."),

    # -- Arctic Tundra (future) --
    _ma("map_arctic_tundra_015",     "animal_polar_bear_043",         8),
    _ma("map_arctic_tundra_015",     "animal_arctic_fox_044",        16),
]

MAP_INDEX = {m["map_id"]: m for m in MAPS}


# =============================================================================
# 7. HunterSkills (data-dictionary — unchanged from v0.1)
# =============================================================================

HUNTER_SKILLS: List[Dict[str, Any]] = [
    dict(skill_id="skill_capture_bonus",
         name_en="Capture Bonus",     name_ja="捕獲補正",
         effect_type="additive_percent",
         effect_min=-20, effect_max=+30,
         description="Added to the base capture success rate. Applied on top of Map difficulty."),
    dict(skill_id="skill_rare_find_bonus",
         name_en="Rare Find Bonus",   name_ja="レア発見補正",
         effect_type="additive_percent",
         effect_min=-20, effect_max=+30,
         description="Bias toward drawing a rarer Animal from the Map's spawn table when a capture succeeds."),
    dict(skill_id="skill_speed_bonus",
         name_en="Speed Bonus",       name_ja="速度補正",
         effect_type="additive_percent",
         effect_min=-20, effect_max=+30,
         description="Reduces the expedition_minutes for the assigned Map by this percentage."),
    dict(skill_id="skill_biome_affinity",
         name_en="Biome Affinity",    name_ja="バイオーム適性",
         effect_type="biome_multiplier",
         effect_min=0, effect_max=0,
         description="Hunter.preferred_biome_id matching Map.biome_id grants ExpeditionRules biome_affinity_bonus."),
]


# =============================================================================
# 8. ExpeditionRules  (v0.2 review decisions applied)
# =============================================================================

EXPEDITION_RULES: List[Dict[str, Any]] = [
    dict(rule_id="expedition_rule_base_success_rate",
         rule_name="Base success rate",
         value=60, unit="percent",
         description="Baseline chance an expedition returns with any capture, before Map difficulty or Hunter skill."),

    dict(rule_id="expedition_rule_minimum_minutes",
         rule_name="Minimum expedition minutes",
         value=5, unit="minutes",
         description="Absolute floor after speed bonuses."),

    dict(rule_id="expedition_rule_maximum_minutes",
         rule_name="Maximum expedition minutes",
         value=1440, unit="minutes",
         description="Absolute ceiling (24 h)."),

    dict(rule_id="expedition_rule_base_capture_attempts",
         rule_name="Base capture attempts per expedition",
         value=1, unit="count",
         description="v0.2: single-attempt per expedition. Multi-attempt logic deferred."),

    dict(rule_id="expedition_rule_failure_penalty_g",
         rule_name="Failure penalty (G)",
         value=0, unit="G",
         description="DECIDED (v0.2): no additional G penalty on failure. Only the sunk dispatch + contract G is lost."),

    dict(rule_id="expedition_rule_release_reward_ratio",
         rule_name="Release reward ratio",
         value=0, unit="ratio",
         description="DECIDED (v0.2): Releasing an Animal returns 0 G. Non-G rewards (Conservation Point etc.) considered later."),

    dict(rule_id="expedition_rule_biome_affinity_bonus",
         rule_name="Biome affinity bonus",
         value=10, unit="percent",
         description="Additive capture bonus when Hunter.preferred_biome_id matches Map.biome_id."),

    dict(rule_id="expedition_rule_biome_mismatch_penalty",
         rule_name="Biome mismatch penalty",
         value=0, unit="percent",
         description="v0.2: no penalty for mismatch. Purely reward the match."),

    dict(rule_id="expedition_rule_hunter_rank_gate_soft",
         rule_name="Hunter rank gate mode",
         value=0, unit="mode (1=hard, 0=soft)",
         description="DECIDED (v0.2): soft. Low-rank Hunter may be dispatched to a Map beyond recommended_hunter_rank — risk is higher, but access is allowed."),

    dict(rule_id="expedition_rule_rare_find_source",
         rule_name="Rare Find source",
         value=1, unit="mode (1=post-capture, 0=pre-capture)",
         description="v0.2: after a successful capture, the game rolls the actual species from the Map's spawn table weighted by rarity + rare_find_bonus."),

    dict(rule_id="expedition_rule_session_short_max_minutes",
         rule_name="Short-session Map ceiling",
         value=30, unit="minutes",
         description="Design intent: Maps with expedition_minutes ≤ 30 are 'short session'."),

    dict(rule_id="expedition_rule_session_medium_max_minutes",
         rule_name="Medium-session Map ceiling",
         value=240, unit="minutes",
         description="Design intent: 31..240 min = 'medium day-time expedition'."),

    dict(rule_id="expedition_rule_session_long_min_minutes",
         rule_name="Long-session Map floor",
         value=360, unit="minutes",
         description="Design intent: ≥ 360 min = 'overnight expedition'."),
]


# =============================================================================
# 9. Review sheet
#
# Every v0.1 entry is either DECIDED (with v0.2 outcome) or DEFERRED.
# Additionally, all v0.2 open questions are added.
# =============================================================================

REVIEW: List[Dict[str, Any]] = [
    # -- Inherited from v0.1 with decisions ---------------------------------
    dict(review_id="review_001", category="Scope",
         question="v0.1: 12 maps + 50 animals + 18 hunters. Right ratio for v0.2?",
         current_proposal="v0.2: 15 maps (9 initial Africa + 6 future_expansion), 50 animals (32 initial_africa / 17 future_region / 1 special_event), 18 hunters (unchanged).",
         reason="Africa-first policy narrows the initial playable set without discarding the future roster.",
         priority="high", decision="REVISED — 9/50/18 initial with the split above", notes=""),

    dict(review_id="review_002", category="Scope",
         question="5 rarity tiers (Common → Legendary), no Mythic — final?",
         current_proposal="Keep 5 tiers.",
         reason="Rarity is now defined as in-game encounter rarity (not IUCN status). Room to add IUCN status later as a separate `conservation_status` column.",
         priority="high", decision="YES — 5 tiers only", notes="Clarified: Rarity ≠ IUCN status. See review_020."),

    dict(review_id="review_003", category="Regions",
         question="v0.1: Africa-heavy but global. Is v0.2 correct in going Africa-only for initial?",
         current_proposal="Africa-only for initial; other continents are future_expansion tags on the Map/Animal rows.",
         reason="Confirmed by human review of v0.1.",
         priority="high", decision="YES — Africa-first", notes=""),

    dict(review_id="review_004", category="Habitat",
         question="v0.1 placed several species outside their real range as game abstractions. Keep or fix?",
         current_proposal="Fix — all game-abstraction MapAnimals rows removed. Affected species are now `availability_phase = future_region` with a placement_note.",
         reason="Real habitat = the game's design constraint (v0.2). 'Access an Animal only by unlocking its real region' is now a progression driver.",
         priority="high", decision="FIX — no more geographic abstractions", notes=""),

    dict(review_id="review_005", category="Hunters",
         question="6-step rank system (Bronze → Master) — right granularity?",
         current_proposal="Retain for v0.2.",
         reason="Not yet exercised by real progression; safe to keep pending v0.3 economy work.",
         priority="medium", decision="TENTATIVE — retain", notes=""),

    dict(review_id="review_006", category="Hunters",
         question="18 hunters — right initial roster size?",
         current_proposal="Retain 18 for v0.2 (unchanged from v0.1).",
         reason="Every archetype has ≥1 example.",
         priority="medium", decision="TENTATIVE — retain", notes=""),

    dict(review_id="review_007", category="Hunters",
         question="Hunter skills as columns on Hunters sheet vs a HunterSkillAssignments join sheet?",
         current_proposal="Retain columns + HunterSkills sheet as a data dictionary.",
         reason="Simpler; every Hunter still has the same 3 numeric skills.",
         priority="medium", decision="YES — columns kept", notes=""),

    dict(review_id="review_008", category="Maps",
         question="Map unlock — was 'minimum Hunter rank owned'. Right approach?",
         current_proposal="NO. Player does not own Hunters (Guild-pool contract model). Replaced by `unlock_rule` + `unlock_value` (draft: zoo_value threshold) + separate `recommended_hunter_rank` (guidance only).",
         reason="Owning a Hunter is not a game concept in v0.2. Progression must be measurable in state the Player actually keeps (Zoo Value, expeditions completed, etc.).",
         priority="high", decision="REVISED — separate unlock_rule from Hunter rank; draft is zoo_value", notes="See review_022 for unlock_rule alternatives."),

    dict(review_id="review_009", category="Expedition",
         question="10 min → 1200 min expedition range — still appropriate?",
         current_proposal="Retain range. Add three named session tiers in ExpeditionRules for clarity: short (≤30m), medium (31..240m), long (≥360m).",
         reason="Matches ADR-0002 §9 and the different lifestyle intent.",
         priority="medium", decision="YES — retain, formalise session tiers", notes=""),

    dict(review_id="review_010", category="Economy",
         question="G cost gradient 50G → 2400G for dispatch. Final?",
         current_proposal="DEFERRED — the daily G income model has not been designed. Numbers stay draft.",
         reason="Cannot balance dispatch cost without knowing G/day yield.",
         priority="high", decision="DEFERRED — pending G economy design (review_023)", notes=""),

    dict(review_id="review_011", category="Economy",
         question="Legendary base ≈ 20-26× Common Zoo Value. Final?",
         current_proposal="DEFERRED. Zoo Value formula (with diminishing returns, ADR-0002 §12) has not been decided.",
         reason="Multiplier only makes sense once the formula around it is fixed.",
         priority="high", decision="DEFERRED — pending Zoo Value formula design (review_024)", notes=""),

    dict(review_id="review_012", category="Presentation",
         question="How should visitor_appeal be used?",
         current_proposal="DEFERRED. Feeds the future Visitor / G formula (ADR-0002 §13).",
         reason="Not consumed by v0.2 gameplay; recorded for the balance conversation.",
         priority="medium", decision="DEFERRED", notes=""),

    dict(review_id="review_013", category="Gameplay",
         question="Capture failure — extra G penalty beyond sunk dispatch cost?",
         current_proposal="NO — DECIDED.",
         reason="ADR-0002 §10 fixes only capture_success / no_capture. Additional penalty is a design choice; v0.2 declines.",
         priority="medium", decision="NO — no extra G penalty", notes=""),

    dict(review_id="review_014", category="Gameplay",
         question="Release — does the Player get any G back?",
         current_proposal="NO — DECIDED.",
         reason="Selling Animals is explicitly not the core loop (ADR-0002 §11). A non-G Conservation reward is left open (review_025).",
         priority="high", decision="NO — 0 G refund", notes="Non-G reward open — review_025."),

    dict(review_id="review_015", category="Guild",
         question="Hunter hire model — one-off contract vs upkeep vs ownership?",
         current_proposal="One-off Guild-pool contract per expedition. No ownership. No upkeep. Column renamed hire_cost_g → contract_cost_g.",
         reason="Confirmed by human review of v0.1.",
         priority="high", decision="YES — one-off contract, no ownership", notes=""),

    dict(review_id="review_016", category="Balance",
         question="Northern White Rhinoceros in v0.1 normal spawn — appropriate?",
         current_proposal="Remove from normal spawn. Tag `availability_phase = special_event`. No MapAnimals row.",
         reason="Only 2 individuals remain in reality (2026). Educational / conservation event surface is more appropriate.",
         priority="high", decision="REMOVE from normal spawn — special_event only", notes=""),

    dict(review_id="review_017", category="Documentation",
         question="Bilingual (JA/EN) column pair — right multilingual approach?",
         current_proposal="Retain.",
         reason="Matches the docs/reports bilingual policy.",
         priority="low", decision="YES — retain", notes=""),

    # -- NEW open questions surfaced by the v0.2 pass ----------------------
    dict(review_id="review_018", category="Regions",
         question="Region unlock order — which continent unlocks first after Africa?",
         current_proposal="Undecided. Draft order for later discussion: South America → Southeast Asia → Oceania → South Asia → Arctic → East Asia → West Africa (adding maps that host the remaining future_region species).",
         reason="Every future_region Animal is blocked until its region ships.",
         priority="high", decision="", notes=""),

    dict(review_id="review_019", category="Regions",
         question="Region unlock granularity — one continent at a time, or one Map at a time inside a continent?",
         current_proposal="Continent-at-a-time as the default. Story-driven Map-at-a-time as a variant for high-cost Maps like Annamite Range.",
         reason="Continent framing gives a clean 'new expansion' narrative; per-Map lets the design team pace prestige Animals.",
         priority="high", decision="", notes=""),

    dict(review_id="review_020", category="Presentation",
         question="Should Animals also carry a conservation_status column (IUCN LC/NT/VU/EN/CR/EW/EX)?",
         current_proposal="Not in v0.2. Rarity is game-facing; conservation status is real-world-facing. Add when the game has a place to surface it.",
         reason="Two axes exist; keeping them separated in the data prevents confusion.",
         priority="medium", decision="", notes=""),

    dict(review_id="review_021", category="Habitat",
         question="Ethiopian Highlands has only Ethiopian Wolf as a signature endemic in v0.2. Should we introduce Gelada / Walia Ibex to give the map real fauna diversity?",
         current_proposal="Deferred to v0.3. Chacma Baboon rows on Ethiopian / Kilimanjaro are flagged needs_review as a genus-level proxy until species are added.",
         reason="v0.2 scope was to reorganise; adding new species is a v0.3 concern.",
         priority="medium", decision="", notes=""),

    dict(review_id="review_022", category="Progression",
         question="Map unlock mechanic — what is the single unlock_rule for v0.2 draft?",
         current_proposal="`zoo_value` threshold per Map (values in Maps sheet). Alternatives: cumulative_expedition_count, distinct_species_captured, seasonal_pass, story_beat.",
         reason="Zoo Value already exists as the primary competitive metric (ADR-0002 §12); reusing it avoids inventing a new counter.",
         priority="high", decision="", notes=""),

    dict(review_id="review_023", category="Economy",
         question="Design the daily G income model — what does a Player earn per real-world day at each Zoo Value?",
         current_proposal="Undecided. Needs: starter G/day, typical active-player G/day, passive Zoo G/day, event G/day, future IAP G exchange scale.",
         reason="Without this, dispatch cost and contract cost cannot be balanced.",
         priority="high", decision="", notes=""),

    dict(review_id="review_024", category="Economy",
         question="Design the Zoo Value formula (with diminishing returns).",
         current_proposal="Undecided. ADR-0002 §12 fixes the diminishing-returns constraint; formula shape (log, tiered, unique-species-multiplier, event-boost) is open.",
         reason="Sets the target function every other economic decision optimises against.",
         priority="high", decision="", notes=""),

    dict(review_id="review_025", category="Gameplay",
         question="Non-G reward on Release — Conservation Points, Reputation, Achievement, Collection progress?",
         current_proposal="Undecided. G is deliberately 0 (review_014); adding a non-G symbolic reward could reinforce the 'not a trading game' framing without breaking it.",
         reason="Otherwise Release feels punitive despite being the intended disposal path.",
         priority="medium", decision="", notes=""),

    dict(review_id="review_026", category="Content",
         question="Future Map release cadence — how often does a new Region ship?",
         current_proposal="Undecided. Candidates: seasonal (quarterly), event-driven, or tied to World-First / achievement completion.",
         reason="Sets the long-term content roadmap and the pace at which future_region Animals become playable.",
         priority="medium", decision="", notes=""),

    dict(review_id="review_027", category="Hunters",
         question="Hunter names review — every Hunter must be a fictional character with no strong overlap with a specific real person.",
         current_proposal="v0.2 renames: Tenzing Sherpa → Nima Kirat; Aiyana Redcloud → Sky Winterhawk; Ravi Chandra → Ravi Menon; Aiko Fujimori → Aiko Tanabe. Others reviewed and considered safely fictional common name combinations.",
         reason="ADR-adjacent — Hunter roster is public-facing content.",
         priority="high", decision="YES — 4 renames applied; remaining names accepted as fictional common combinations", notes=""),

    dict(review_id="review_028", category="Data",
         question="Introduce Biomes as a shared master with FK-checked references from Maps / Animals / Hunters?",
         current_proposal="YES — new Biomes sheet with 6 biomes. Maps.biome_id, Animals.habitat_biome_id, Hunters.preferred_biome_id all validated against it (Hunters may use the literal 'any').",
         reason="v0.1 used free-text biome labels; scaling to more content is fragile without a controlled vocabulary.",
         priority="high", decision="YES — Biomes master added in v0.2", notes=""),

    dict(review_id="review_029", category="Rules",
         question="Hunter rank gate on Maps — hard or soft?",
         current_proposal="SOFT. Any Hunter may accept any unlocked Map; below-recommended-rank just means lower success rate. See ExpeditionRules.expedition_rule_hunter_rank_gate_soft = 0.",
         reason="Matches the design principle 'access permitted, risk high' from the v0.2 spec.",
         priority="high", decision="SOFT gate", notes=""),
]


# =============================================================================
# Excel writing (shared with v0.1)
# =============================================================================

HEADER_FILL = PatternFill(start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFFFF")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _write_sheet(ws, header: List[str], rows: List[Dict[str, Any]], widths: Dict[str, int]) -> None:
    for col_idx, key in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx, value=key)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, key in enumerate(header, start=1):
            value = row.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = LEFT
    for col_idx, key in enumerate(header, start=1):
        width = widths.get(key, 18)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    last_col_letter = get_column_letter(len(header))
    ws.auto_filter.ref = f"A1:{last_col_letter}{max(len(rows) + 1, 2)}"


# =============================================================================
# Validation (v0.2 strengthened)
# =============================================================================

def validate() -> Tuple[bool, List[str]]:
    problems: List[str] = []

    # -- Unique IDs per sheet ---------------------------------------------
    for name, rows, key in [
        ("Biomes",           BIOMES,           "biome_id"),
        ("Rarities",         RARITIES,         "rarity_id"),
        ("Maps",             MAPS,             "map_id"),
        ("Animals",          ANIMALS,          "animal_id"),
        ("Hunters",          HUNTERS,          "hunter_id"),
        ("HunterSkills",     HUNTER_SKILLS,    "skill_id"),
        ("ExpeditionRules",  EXPEDITION_RULES, "rule_id"),
        ("Review",           REVIEW,           "review_id"),
    ]:
        seen = set()
        for r in rows:
            v = r[key]
            if v in seen:
                problems.append(f"{name}: duplicate {key} = {v!r}")
            seen.add(v)

    # -- FK checks --------------------------------------------------------
    map_ids     = {m["map_id"] for m in MAPS}
    animal_ids  = {a["animal_id"] for a in ANIMALS}
    rarity_ids  = {r["rarity_id"] for r in RARITIES}
    biome_ids   = {b["biome_id"] for b in BIOMES}

    for m in MAPS:
        if m["biome_id"] not in biome_ids:
            problems.append(f"Maps.{m['map_id']}: biome_id {m['biome_id']!r} not in Biomes")

    for a in ANIMALS:
        if a["habitat_biome_id"] not in biome_ids:
            problems.append(f"Animals.{a['animal_id']}: habitat_biome_id {a['habitat_biome_id']!r} not in Biomes")
        if a["rarity_id"] not in rarity_ids:
            problems.append(f"Animals.{a['animal_id']}: rarity_id {a['rarity_id']!r} not in Rarities")

    for h in HUNTERS:
        pb = h["preferred_biome_id"]
        if pb != "any" and pb not in biome_ids:
            problems.append(f"Hunters.{h['hunter_id']}: preferred_biome_id {pb!r} not in Biomes and not 'any'")

    for ma in MAP_ANIMALS_RAW:
        if ma["map_id"] not in map_ids:
            problems.append(f"MapAnimals: map_id {ma['map_id']!r} not in Maps")
        if ma["animal_id"] not in animal_ids:
            problems.append(f"MapAnimals: animal_id {ma['animal_id']!r} not in Animals")

    # -- v0.2 invariants --------------------------------------------------

    # No MapAnimals row for special_event Animals
    for ma in MAP_ANIMALS_RAW:
        a = ANIMAL_INDEX.get(ma["animal_id"])
        if a and a["availability_phase"] == "special_event":
            problems.append(
                f"MapAnimals: special_event Animal {a['animal_id']!r} must not have a spawn row "
                f"(found on {ma['map_id']!r})"
            )

    # No future_region Animal on an initial_africa Map
    for ma in MAP_ANIMALS_RAW:
        a = ANIMAL_INDEX.get(ma["animal_id"])
        m = MAP_INDEX.get(ma["map_id"])
        if a and m and a["availability_phase"] == "future_region" and m["availability_phase"] == "initial_africa":
            problems.append(
                f"MapAnimals: future_region Animal {a['animal_id']!r} placed on initial_africa Map "
                f"{m['map_id']!r} (blocked by v0.2 real-habitat rule)"
            )

    # Northern White Rhinoceros must NOT appear in MapAnimals
    for ma in MAP_ANIMALS_RAW:
        if ma["animal_id"] == "animal_northern_white_rhino_039":
            problems.append(
                "MapAnimals: Northern White Rhinoceros must not have any spawn row "
                "(availability_phase = special_event)"
            )

    # Every initial_africa Animal has at least one MapAnimals row
    initial_animals = {a["animal_id"] for a in ANIMALS if a["availability_phase"] == "initial_africa"}
    used = {ma["animal_id"] for ma in MAP_ANIMALS_RAW}
    for aid in sorted(initial_animals):
        if aid not in used:
            problems.append(f"initial_africa Animal {aid!r} has no MapAnimals row")

    # Every initial_africa Map has at least one MapAnimals row
    initial_maps = {m["map_id"] for m in MAPS if m["availability_phase"] == "initial_africa"}
    used_maps = {ma["map_id"] for ma in MAP_ANIMALS_RAW}
    for mid in sorted(initial_maps):
        if mid not in used_maps:
            problems.append(f"initial_africa Map {mid!r} has no MapAnimals row")

    # future_region Animals should ideally have a matching future_expansion Map row
    future_animals = {a["animal_id"] for a in ANIMALS if a["availability_phase"] == "future_region"}
    future_maps = {m["map_id"] for m in MAPS if m["availability_phase"] == "future_expansion"}
    for aid in sorted(future_animals):
        rows_for_a = [ma for ma in MAP_ANIMALS_RAW if ma["animal_id"] == aid]
        if not rows_for_a:
            # Orphan future_region — legal but flag as notice
            a = ANIMAL_INDEX[aid]
            problems.append(f"NOTICE: future_region Animal {aid!r} has no MapAnimals row (placement_note: {a.get('placement_note','')!r})")
        else:
            for ma in rows_for_a:
                m = MAP_INDEX.get(ma["map_id"])
                if m and m["availability_phase"] != "future_expansion":
                    problems.append(f"MapAnimals: future_region Animal {aid!r} placed on non-future Map {ma['map_id']!r}")

    # -- Required non-null fields -----------------------------------------
    def _require(name: str, rows: List[Dict[str, Any]], fields: List[str]) -> None:
        for r in rows:
            for f in fields:
                v = r.get(f)
                if v is None or (isinstance(v, str) and not v.strip()):
                    ident = r.get("name_en") or r.get("hunter_id") or r.get("rule_id") or r.get("review_id") or "?"
                    problems.append(f"{name}: {ident} — empty {f}")

    _require("Biomes",          BIOMES,          ["biome_id", "name_en", "name_ja"])
    _require("Rarities",        RARITIES,        ["rarity_id", "name_en", "name_ja",
                                                  "sort_order", "base_multiplier"])
    _require("Maps",            MAPS,            ["map_id", "name_en", "name_ja",
                                                  "region", "biome_id",
                                                  "availability_phase", "unlock_rule",
                                                  "difficulty",
                                                  "expedition_minutes", "base_cost_g"])
    _require("Animals",         ANIMALS,         ["animal_id", "name_en", "name_ja",
                                                  "rarity_id", "habitat_biome_id",
                                                  "availability_phase",
                                                  "base_zoo_value", "capture_difficulty"])
    _require("Hunters",         HUNTERS,         ["hunter_id", "name", "rank",
                                                  "specialty", "preferred_biome_id",
                                                  "contract_cost_g"])
    _require("HunterSkills",    HUNTER_SKILLS,   ["skill_id", "name_en", "effect_type"])
    _require("ExpeditionRules", EXPEDITION_RULES,["rule_id", "rule_name", "value", "unit"])
    _require("Review",          REVIEW,          ["review_id", "category", "question",
                                                  "current_proposal", "priority"])

    # -- Numeric range sanity ---------------------------------------------
    for m in MAPS:
        if not (1 <= m["difficulty"] <= 5):
            problems.append(f"Maps.{m['map_id']}: difficulty out of range")
        if not (0 <= m["recommended_hunter_rank"] <= 6):
            problems.append(f"Maps.{m['map_id']}: recommended_hunter_rank out of range")
        if m["availability_phase"] not in ("initial_africa", "future_expansion"):
            problems.append(f"Maps.{m['map_id']}: availability_phase invalid")

    for a in ANIMALS:
        if not (1 <= a["capture_difficulty"] <= 5):
            problems.append(f"Animals.{a['animal_id']}: capture_difficulty out of range")
        if not (1 <= a["visitor_appeal"] <= 100):
            problems.append(f"Animals.{a['animal_id']}: visitor_appeal out of 1..100")
        if a["availability_phase"] not in ("initial_africa", "future_region", "special_event"):
            problems.append(f"Animals.{a['animal_id']}: availability_phase invalid")

    hard_failures = [p for p in problems if not p.startswith("NOTICE")]
    return (len(hard_failures) == 0, problems)


# =============================================================================
# Main
# =============================================================================

def main() -> int:
    ok, problems = validate()
    for p in problems:
        print(("[!] " if not p.startswith("NOTICE") else "[i] ") + p)
    if not ok:
        print("\nvalidation failed — refusing to write xlsx")
        return 1

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Biomes"
    _write_sheet(ws0,
                 header=["biome_id", "name_ja", "name_en", "description_ja", "description_en"],
                 rows=BIOMES,
                 widths={"biome_id": 18, "name_ja": 16, "name_en": 16,
                         "description_ja": 55, "description_en": 60})

    _write_sheet(wb.create_sheet("Rarities"),
                 header=["rarity_id", "name_en", "name_ja", "sort_order",
                         "base_multiplier", "description"],
                 rows=RARITIES,
                 widths={"rarity_id": 22, "name_en": 14, "name_ja": 14,
                         "sort_order": 12, "base_multiplier": 16,
                         "description": 65})

    _write_sheet(wb.create_sheet("Maps"),
                 header=["map_id", "name_ja", "name_en", "region", "biome_id",
                         "availability_phase",
                         "unlock_rule", "unlock_value",
                         "recommended_hunter_rank", "minimum_hunter_rank_gate",
                         "difficulty", "expedition_minutes", "base_cost_g",
                         "risk_level",
                         "description_ja", "description_en"],
                 rows=MAPS,
                 widths={"map_id": 32, "name_ja": 22, "name_en": 22, "region": 16,
                         "biome_id": 16, "availability_phase": 18,
                         "unlock_rule": 18, "unlock_value": 12,
                         "recommended_hunter_rank": 24, "minimum_hunter_rank_gate": 24,
                         "difficulty": 11, "expedition_minutes": 18,
                         "base_cost_g": 14, "risk_level": 11,
                         "description_ja": 55, "description_en": 55})

    _write_sheet(wb.create_sheet("Animals"),
                 header=["animal_id", "name_ja", "name_en", "category",
                         "rarity_id", "availability_phase", "placement_note",
                         "base_zoo_value", "capture_difficulty",
                         "growth_rate", "visitor_appeal",
                         "habitat_biome_id", "size", "active_time",
                         "description_ja", "description_en"],
                 rows=ANIMALS,
                 widths={"animal_id": 36, "name_ja": 22, "name_en": 24, "category": 14,
                         "rarity_id": 18, "availability_phase": 18, "placement_note": 55,
                         "base_zoo_value": 14, "capture_difficulty": 18,
                         "growth_rate": 12, "visitor_appeal": 14,
                         "habitat_biome_id": 18, "size": 10, "active_time": 12,
                         "description_ja": 55, "description_en": 55})

    _write_sheet(wb.create_sheet("Hunters"),
                 header=["hunter_id", "name", "rank", "level", "specialty",
                         "preferred_biome_id", "capture_bonus", "rare_find_bonus",
                         "speed_bonus", "contract_cost_g",
                         "personality", "description"],
                 rows=HUNTERS,
                 widths={"hunter_id": 32, "name": 22, "rank": 12, "level": 8,
                         "specialty": 24, "preferred_biome_id": 18,
                         "capture_bonus": 14, "rare_find_bonus": 16,
                         "speed_bonus": 12, "contract_cost_g": 16,
                         "personality": 45, "description": 55})

    map_animal_rows = []
    for i, ma in enumerate(MAP_ANIMALS_RAW, start=1):
        row = {"map_animal_id": f"map_animal_{i:03d}"}
        row.update(ma)
        map_animal_rows.append(row)
    _write_sheet(wb.create_sheet("MapAnimals"),
                 header=["map_animal_id", "map_id", "animal_id", "spawn_weight",
                         "capture_modifier", "needs_review", "notes"],
                 rows=map_animal_rows,
                 widths={"map_animal_id": 18, "map_id": 32, "animal_id": 36,
                         "spawn_weight": 14, "capture_modifier": 16,
                         "needs_review": 14, "notes": 65})

    _write_sheet(wb.create_sheet("HunterSkills"),
                 header=["skill_id", "name_ja", "name_en", "effect_type",
                         "effect_min", "effect_max", "description"],
                 rows=HUNTER_SKILLS,
                 widths={"skill_id": 24, "name_ja": 20, "name_en": 22,
                         "effect_type": 22, "effect_min": 12, "effect_max": 12,
                         "description": 65})

    _write_sheet(wb.create_sheet("ExpeditionRules"),
                 header=["rule_id", "rule_name", "value", "unit", "description"],
                 rows=EXPEDITION_RULES,
                 widths={"rule_id": 44, "rule_name": 34, "value": 10,
                         "unit": 24, "description": 65})

    _write_sheet(wb.create_sheet("Review"),
                 header=["review_id", "category", "question", "current_proposal",
                         "reason", "priority", "decision", "notes"],
                 rows=REVIEW,
                 widths={"review_id": 14, "category": 14, "question": 50,
                         "current_proposal": 55, "reason": 55,
                         "priority": 10, "decision": 45, "notes": 30})

    out = Path(__file__).parent / "WildLive-Game-Master-Draft-v0.2.xlsx"
    _deterministic_save(wb, out)
    print(f"wrote {out}")
    print(f"    sheets:          {len(wb.sheetnames)}  ({', '.join(wb.sheetnames)})")
    print(f"    Biomes:          {len(BIOMES)}")
    print(f"    Rarities:        {len(RARITIES)}")

    initial_maps = [m for m in MAPS if m["availability_phase"] == "initial_africa"]
    future_maps  = [m for m in MAPS if m["availability_phase"] == "future_expansion"]
    initial_a    = [a for a in ANIMALS if a["availability_phase"] == "initial_africa"]
    future_a     = [a for a in ANIMALS if a["availability_phase"] == "future_region"]
    special_a    = [a for a in ANIMALS if a["availability_phase"] == "special_event"]

    print(f"    Maps:            {len(MAPS)}  (initial_africa {len(initial_maps)} / future_expansion {len(future_maps)})")
    print(f"    Animals:         {len(ANIMALS)}  (initial_africa {len(initial_a)} / future_region {len(future_a)} / special_event {len(special_a)})")
    print(f"    Hunters:         {len(HUNTERS)}")
    print(f"    MapAnimals:      {len(MAP_ANIMALS_RAW)}")
    print(f"    HunterSkills:    {len(HUNTER_SKILLS)}")
    print(f"    ExpeditionRules: {len(EXPEDITION_RULES)}")
    print(f"    Review:          {len(REVIEW)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
