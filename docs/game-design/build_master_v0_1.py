#!/usr/bin/env python3
"""Build WildLive-Game-Master-Draft-v0.1.xlsx.

This script IS the source of truth for the draft master data. When a
human edits the Excel file directly, that edit lives only in the Excel
file until the human (or an AI on human instruction) mirrors it back
into this script. Re-running the script regenerates the Excel from
these Python lists.

Design constraints inherited from docs/adr/0002-game-system-foundation.md:

  - Real wildlife only. No fantasy species.
  - "African-inspired working assumption" for regions is honoured
    (7 of 12 maps are African) while still allowing global species
    variety.
  - Species rarity + individual rarity are separate concerns
    (individual rarity handled by future traits, not modelled here).
  - No power inflation — new content is rarer, not stronger.
  - Every value in this draft is a proposal for human review. Nothing
    here is a final balance decision.

Run:

    python3 docs/game-design/build_master_v0_1.py

Output:

    docs/game-design/WildLive-Game-Master-Draft-v0.1.xlsx
"""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# =============================================================================
# 1. Rarities
# =============================================================================

RARITIES: List[Dict[str, Any]] = [
    # rarity_id, name_en, name_ja, sort_order, base_multiplier, description
    dict(rarity_id="rarity_common",    name_en="Common",    name_ja="コモン",     sort_order=1, base_multiplier=1.0,
         description="Widely distributed. Encountered on most easy expeditions."),
    dict(rarity_id="rarity_uncommon",  name_en="Uncommon",  name_ja="アンコモン", sort_order=2, base_multiplier=1.5,
         description="Regionally common but not everywhere. Reliable mid-tier."),
    dict(rarity_id="rarity_rare",      name_en="Rare",      name_ja="レア",       sort_order=3, base_multiplier=2.5,
         description="Iconic species that most players will eventually collect."),
    dict(rarity_id="rarity_epic",      name_en="Epic",      name_ja="エピック",   sort_order=4, base_multiplier=5.0,
         description="Critically endangered or otherwise very hard to find. Prestige tier."),
    dict(rarity_id="rarity_legendary", name_en="Legendary", name_ja="レジェンド", sort_order=5, base_multiplier=12.0,
         description="Near-mythical real animals. A handful per player expected in a long life."),
]


# =============================================================================
# 2. Maps  (12 maps — 7 African, 5 global)
#
# "region"  = broad geographic label (informational)
# "biome"   = gameplay-relevant biome tag used by MapAnimals + Hunter affinity
# "difficulty" = 1..5 (Draft scale — see Review sheet)
# "unlock_rank" = minimum Hunter rank required to accept a mission here
#                 (0 = always available, 1 = Bronze, 2 = Silver, …)
# "expedition_minutes" = base wait; final wait scales with Hunter speed
#                        (bounded by ExpeditionRules min/max)
# "base_cost_g" = G paid to dispatch (before Hunter contract cost)
# "risk_level"  = 1..5 (Draft scale — chance of no-capture rises with this)
# =============================================================================

MAPS: List[Dict[str, Any]] = [
    dict(map_id="map_kenyan_savanna_001",
         name_en="Kenyan Savanna",       name_ja="ケニアのサバンナ",
         region="East Africa", biome="savanna", difficulty=1, unlock_rank=0,
         expedition_minutes=10,  base_cost_g=50,   risk_level=1,
         description_en="Wide-open grassland teeming with iconic African wildlife. First place every player starts.",
         description_ja="広大な草原に象徴的なアフリカの動物が集う地。すべてのプレイヤーが最初に訪れる場所。"),

    dict(map_id="map_serengeti_plains_002",
         name_en="Serengeti Plains",     name_ja="セレンゲティ平原",
         region="East Africa", biome="savanna", difficulty=2, unlock_rank=0,
         expedition_minutes=20,  base_cost_g=90,   risk_level=1,
         description_en="Endless migration corridor. Larger herds, but rangier terrain than a starter savanna.",
         description_ja="果てしない移動経路。群れは大きいが起伏に富み、初心者向けの草原よりやや荒々しい。"),

    dict(map_id="map_okavango_delta_003",
         name_en="Okavango Delta",       name_ja="オカバンゴ湿地",
         region="Southern Africa", biome="wetland", difficulty=2, unlock_rank=1,
         expedition_minutes=40,  base_cost_g=180,  risk_level=2,
         description_en="Seasonal wetland maze. Water-adapted species dominate; navigation is slow.",
         description_ja="季節性の湿地帯。水辺に適応した動物が主役で、移動は遅く難しい。"),

    dict(map_id="map_namib_desert_004",
         name_en="Namib Desert",         name_ja="ナミブ砂漠",
         region="Southern Africa", biome="desert", difficulty=3, unlock_rank=1,
         expedition_minutes=60,  base_cost_g=240,  risk_level=3,
         description_en="Extreme aridity. Few species, but the ones that live here are specialised and elusive.",
         description_ja="極度の乾燥地帯。生息種は少ないが、その一つひとつが特殊化していて捕獲は難しい。"),

    dict(map_id="map_atlas_mountains_005",
         name_en="Atlas Mountains",      name_ja="アトラス山脈",
         region="North Africa", biome="mountain", difficulty=3, unlock_rank=2,
         expedition_minutes=120, base_cost_g=360,  risk_level=3,
         description_en="High-altitude escarpments. Cold nights and thin trails favour mountain specialists.",
         description_ja="高地の急峻な断崖。冷え込む夜と細い道は山岳スペシャリスト向き。"),

    dict(map_id="map_kilimanjaro_slopes_006",
         name_en="Kilimanjaro Slopes",   name_ja="キリマンジャロ斜面",
         region="East Africa", biome="mountain", difficulty=4, unlock_rank=2,
         expedition_minutes=240, base_cost_g=600,  risk_level=3,
         description_en="Alpine climb from rainforest to snowline. Multiple biomes in one expedition; long return.",
         description_ja="熱帯雨林から雪線までを登る長い遠征。複数のバイオームを跨ぐが往復時間も長い。"),

    dict(map_id="map_congo_rainforest_007",
         name_en="Congo Rainforest",     name_ja="コンゴ熱帯雨林",
         region="Central Africa", biome="rainforest", difficulty=4, unlock_rank=3,
         expedition_minutes=360, base_cost_g=900,  risk_level=4,
         description_en="Dense equatorial forest. Great apes and elusive antelopes; almost no line-of-sight.",
         description_ja="赤道直下の密林。大型類人猿と幻の偶蹄類。見通しはほぼ利かない。"),

    dict(map_id="map_amazon_rainforest_008",
         name_en="Amazon Rainforest",    name_ja="アマゾン熱帯雨林",
         region="South America", biome="rainforest", difficulty=4, unlock_rank=3,
         expedition_minutes=420, base_cost_g=1100, risk_level=4,
         description_en="World's largest rainforest. Apex predators and enormous prey — but the canopy hides them.",
         description_ja="世界最大の熱帯雨林。頂点捕食者と巨大な獲物が林冠に潜む。"),

    dict(map_id="map_borneo_jungle_009",
         name_en="Borneo Jungle",        name_ja="ボルネオのジャングル",
         region="Southeast Asia", biome="rainforest", difficulty=4, unlock_rank=3,
         expedition_minutes=480, base_cost_g=1200, risk_level=4,
         description_en="Old-growth tropical forest. Home to red apes and the rarest big cats in the game.",
         description_ja="原生の熱帯林。赤い類人猿と、本作屈指の希少大型ネコ科動物が住む。"),

    dict(map_id="map_australian_outback_010",
         name_en="Australian Outback",   name_ja="オーストラリア奥地",
         region="Oceania", biome="desert", difficulty=3, unlock_rank=2,
         expedition_minutes=180, base_cost_g=480,  risk_level=3,
         description_en="Vast red interior. Wildlife found nowhere else on the planet.",
         description_ja="広大な赤い内陸。地球上のここにしか棲まない野生動物が集う。"),

    dict(map_id="map_himalayan_foothills_011",
         name_en="Himalayan Foothills",  name_ja="ヒマラヤ山麓",
         region="South Asia", biome="mountain", difficulty=5, unlock_rank=4,
         expedition_minutes=600, base_cost_g=1800, risk_level=4,
         description_en="Very high altitude. Only elite Hunters return with anything, and even they often don't.",
         description_ja="極めて高い高度。エリートハンターだけが何かを持ち帰り、それでも空手ばしばしば。"),

    dict(map_id="map_arctic_tundra_012",
         name_en="Arctic Tundra",        name_ja="北極ツンドラ",
         region="Arctic", biome="tundra", difficulty=5, unlock_rank=4,
         expedition_minutes=1200, base_cost_g=2400, risk_level=5,
         description_en="Cold, empty, unforgiving. The longest single expedition in the game (~20 hours).",
         description_ja="寒く、空虚で、容赦ない。本作最長級の単発遠征(約20時間)。"),
]


# =============================================================================
# 3. Animals  (50 real species — spread across 5 rarity tiers)
#
# All are real, extant (or recently living) species.
#
# "category"        = big-cat / hoofed / primate / carnivore / marsupial /
#                     reptile / marine / bird / small-mammal
# "base_zoo_value"  = draft base contribution to Zoo Value at Common tier;
#                     final value = base_zoo_value * rarity.base_multiplier
# "capture_difficulty" = 1..5 (per-species base difficulty; final rate also
#                              depends on Map difficulty and Hunter skill)
# "growth_rate"     = 1..5 (how fast a captured Animal's Visitor appeal
#                           ramps up in the Zoo — placeholder for future
#                           in-zoo mechanic; not used by MVP)
# "visitor_appeal"  = 1..100 (draft — how much a single specimen draws
#                             visitors relative to a boring Common one)
# "habitat"         = free-text natural biome (matches Maps.biome loosely)
# "size"            = tiny / small / medium / large / huge
# "active_time"     = diurnal / nocturnal / crepuscular / cathemeral
# =============================================================================

ANIMALS: List[Dict[str, Any]] = [
    # ------ Common (14) --------------------------------------------------
    dict(animal_id="animal_impala_001",             name_en="Impala",              name_ja="インパラ",
         category="hoofed",     rarity_id="rarity_common",   base_zoo_value=10,  capture_difficulty=1, growth_rate=3, visitor_appeal=8,
         habitat="savanna",     size="medium", active_time="diurnal",
         description_en="Graceful savanna antelope. Everywhere on the plains.",
         description_ja="優雅なサバンナのアンテロープ。平原のどこにでもいる。"),

    dict(animal_id="animal_common_zebra_002",       name_en="Common Zebra",        name_ja="サバンナシマウマ",
         category="hoofed",     rarity_id="rarity_common",   base_zoo_value=12,  capture_difficulty=1, growth_rate=3, visitor_appeal=15,
         habitat="savanna",     size="large", active_time="diurnal",
         description_en="Iconic striped equid of the East African plains.",
         description_ja="東アフリカ平原を象徴する縞模様の草食動物。"),

    dict(animal_id="animal_blue_wildebeest_003",    name_en="Blue Wildebeest",     name_ja="オグロヌー",
         category="hoofed",     rarity_id="rarity_common",   base_zoo_value=11,  capture_difficulty=1, growth_rate=3, visitor_appeal=10,
         habitat="savanna",     size="large", active_time="diurnal",
         description_en="The great migrator. Millions cross the Mara every year.",
         description_ja="大移動の主役。年に数百万頭がマラ川を渡る。"),

    dict(animal_id="animal_warthog_004",            name_en="Warthog",             name_ja="イボイノシシ",
         category="hoofed",     rarity_id="rarity_common",   base_zoo_value=10,  capture_difficulty=1, growth_rate=3, visitor_appeal=12,
         habitat="savanna",     size="medium", active_time="diurnal",
         description_en="Tusked pig of the savanna; runs with its tail up.",
         description_ja="サバンナに棲む牙のあるイノシシ。尾をピンと立てて走る。"),

    dict(animal_id="animal_cape_buffalo_005",       name_en="Cape Buffalo",        name_ja="アフリカスイギュウ",
         category="hoofed",     rarity_id="rarity_common",   base_zoo_value=15,  capture_difficulty=2, growth_rate=3, visitor_appeal=14,
         habitat="savanna",     size="large", active_time="diurnal",
         description_en="Heavy, unpredictable herd animal. Feared even by lions.",
         description_ja="重量級で予測不能な群れ動物。ライオンさえ恐れる。"),

    dict(animal_id="animal_chacma_baboon_006",      name_en="Chacma Baboon",       name_ja="チャクマヒヒ",
         category="primate",    rarity_id="rarity_common",   base_zoo_value=10,  capture_difficulty=1, growth_rate=3, visitor_appeal=13,
         habitat="savanna",     size="medium", active_time="diurnal",
         description_en="Large ground-dwelling monkey. Highly social.",
         description_ja="大型の地上性ザル。強く群れて社会を成す。"),

    dict(animal_id="animal_vervet_monkey_007",      name_en="Vervet Monkey",       name_ja="ベルベットモンキー",
         category="primate",    rarity_id="rarity_common",   base_zoo_value=8,   capture_difficulty=1, growth_rate=3, visitor_appeal=14,
         habitat="savanna",     size="small", active_time="diurnal",
         description_en="Small, quick, and inquisitive.",
         description_ja="小柄で敏捷、好奇心旺盛。"),

    dict(animal_id="animal_springbok_008",          name_en="Springbok",           name_ja="スプリングボック",
         category="hoofed",     rarity_id="rarity_common",   base_zoo_value=10,  capture_difficulty=1, growth_rate=3, visitor_appeal=10,
         habitat="savanna",     size="medium", active_time="diurnal",
         description_en="South African antelope famous for its vertical pronking.",
         description_ja="真上に飛び跳ねる姿で有名な南アフリカのアンテロープ。"),

    dict(animal_id="animal_meerkat_009",            name_en="Meerkat",             name_ja="ミーアキャット",
         category="small-mammal", rarity_id="rarity_common", base_zoo_value=10,  capture_difficulty=1, growth_rate=3, visitor_appeal=25,
         habitat="desert",      size="small", active_time="diurnal",
         description_en="Sentry-standing desert mongoose. Popular with visitors.",
         description_ja="立哨する砂漠のマングース。来園者に人気。"),

    dict(animal_id="animal_aardvark_010",           name_en="Aardvark",            name_ja="ツチブタ",
         category="small-mammal", rarity_id="rarity_common", base_zoo_value=12,  capture_difficulty=2, growth_rate=2, visitor_appeal=18,
         habitat="savanna",     size="medium", active_time="nocturnal",
         description_en="Termite-eating burrower. Rarely seen despite being common.",
         description_ja="シロアリを食べる穴掘り屋。数は多いが姿を見せない。"),

    dict(animal_id="animal_hippopotamus_011",       name_en="Common Hippopotamus", name_ja="カバ",
         category="hoofed",     rarity_id="rarity_common",   base_zoo_value=14,  capture_difficulty=2, growth_rate=3, visitor_appeal=20,
         habitat="wetland",     size="huge", active_time="cathemeral",
         description_en="Massive semi-aquatic. One of Africa's deadliest animals.",
         description_ja="半水生の巨体。アフリカで最も危険な動物のひとつ。"),

    dict(animal_id="animal_nile_crocodile_012",     name_en="Nile Crocodile",      name_ja="ナイルワニ",
         category="reptile",    rarity_id="rarity_common",   base_zoo_value=13,  capture_difficulty=2, growth_rate=2, visitor_appeal=15,
         habitat="wetland",     size="large", active_time="diurnal",
         description_en="Ambush predator of every African river.",
         description_ja="アフリカの河川すべてに潜む待ち伏せ捕食者。"),

    dict(animal_id="animal_fennec_fox_013",         name_en="Fennec Fox",          name_ja="フェネックギツネ",
         category="small-mammal", rarity_id="rarity_common", base_zoo_value=10,  capture_difficulty=1, growth_rate=3, visitor_appeal=30,
         habitat="desert",      size="tiny", active_time="nocturnal",
         description_en="Tiny desert fox with huge ears. Widely adored.",
         description_ja="巨大な耳を持つ小さな砂漠のキツネ。人気者。"),

    dict(animal_id="animal_dingo_014",              name_en="Dingo",               name_ja="ディンゴ",
         category="carnivore",  rarity_id="rarity_common",   base_zoo_value=11,  capture_difficulty=2, growth_rate=3, visitor_appeal=14,
         habitat="desert",      size="medium", active_time="cathemeral",
         description_en="Australia's wild dog. Long-legged and wary.",
         description_ja="オーストラリアの野犬。脚が長く用心深い。"),

    # ------ Uncommon (14) ------------------------------------------------
    dict(animal_id="animal_masai_giraffe_015",      name_en="Masai Giraffe",       name_ja="マサイキリン",
         category="hoofed",     rarity_id="rarity_uncommon", base_zoo_value=22,  capture_difficulty=2, growth_rate=2, visitor_appeal=40,
         habitat="savanna",     size="huge", active_time="diurnal",
         description_en="Tallest land animal. Every visitor wants a photo.",
         description_ja="陸上最高身長の動物。来園者は必ず撮影する。"),

    dict(animal_id="animal_african_elephant_016",   name_en="African Elephant",    name_ja="アフリカゾウ",
         category="hoofed",     rarity_id="rarity_uncommon", base_zoo_value=28,  capture_difficulty=3, growth_rate=1, visitor_appeal=50,
         habitat="savanna",     size="huge", active_time="cathemeral",
         description_en="Largest land animal. Slow to age up but hugely valuable.",
         description_ja="陸上最大の動物。成長は遅いが極めて価値が高い。"),

    dict(animal_id="animal_leopard_017",            name_en="Leopard",             name_ja="ヒョウ",
         category="big-cat",    rarity_id="rarity_uncommon", base_zoo_value=25,  capture_difficulty=3, growth_rate=3, visitor_appeal=35,
         habitat="savanna",     size="medium", active_time="nocturnal",
         description_en="Solitary big cat that adapts to every biome from desert to forest.",
         description_ja="どのバイオームにも適応する単独行動の大型ネコ科動物。"),

    dict(animal_id="animal_cheetah_018",            name_en="Cheetah",             name_ja="チーター",
         category="big-cat",    rarity_id="rarity_uncommon", base_zoo_value=24,  capture_difficulty=3, growth_rate=3, visitor_appeal=38,
         habitat="savanna",     size="medium", active_time="diurnal",
         description_en="Fastest land animal. Fragile compared to other big cats.",
         description_ja="陸上最速の動物。他の大型ネコ科より繊細。"),

    dict(animal_id="animal_spotted_hyena_019",      name_en="Spotted Hyena",       name_ja="ブチハイエナ",
         category="carnivore",  rarity_id="rarity_uncommon", base_zoo_value=20,  capture_difficulty=2, growth_rate=3, visitor_appeal=18,
         habitat="savanna",     size="medium", active_time="nocturnal",
         description_en="Highly social carnivore. Clan-based societies.",
         description_ja="高度に社会的な捕食者。氏族制の群れをつくる。"),

    dict(animal_id="animal_african_wild_dog_020",   name_en="African Wild Dog",    name_ja="リカオン",
         category="carnivore",  rarity_id="rarity_uncommon", base_zoo_value=22,  capture_difficulty=3, growth_rate=3, visitor_appeal=22,
         habitat="savanna",     size="medium", active_time="diurnal",
         description_en="Endangered pack hunter with the highest hunting success rate on the continent.",
         description_ja="絶滅危惧。大陸屈指の狩猟成功率を誇る群れハンター。"),

    dict(animal_id="animal_serval_021",             name_en="Serval",              name_ja="サーバル",
         category="small-mammal", rarity_id="rarity_uncommon", base_zoo_value=18, capture_difficulty=2, growth_rate=3, visitor_appeal=28,
         habitat="savanna",     size="small", active_time="nocturnal",
         description_en="Long-legged small cat. Leaps to catch birds.",
         description_ja="脚の長い小型ネコ科。跳躍して鳥を捕える。"),

    dict(animal_id="animal_caracal_022",            name_en="Caracal",             name_ja="カラカル",
         category="small-mammal", rarity_id="rarity_uncommon", base_zoo_value=18, capture_difficulty=2, growth_rate=3, visitor_appeal=26,
         habitat="savanna",     size="small", active_time="nocturnal",
         description_en="Tufted-eared cat of arid and mountainous zones.",
         description_ja="房耳のネコ科。乾燥地帯と山岳に生息。"),

    dict(animal_id="animal_red_kangaroo_023",       name_en="Red Kangaroo",        name_ja="アカカンガルー",
         category="marsupial",  rarity_id="rarity_uncommon", base_zoo_value=20,  capture_difficulty=2, growth_rate=3, visitor_appeal=32,
         habitat="desert",      size="large", active_time="crepuscular",
         description_en="World's largest marsupial. Australia only.",
         description_ja="世界最大の有袋類。オーストラリアのみ。"),

    dict(animal_id="animal_ethiopian_wolf_024",     name_en="Ethiopian Wolf",      name_ja="エチオピアオオカミ",
         category="carnivore",  rarity_id="rarity_uncommon", base_zoo_value=22,  capture_difficulty=3, growth_rate=3, visitor_appeal=25,
         habitat="mountain",    size="medium", active_time="diurnal",
         description_en="Endangered highland canid. Only ~500 left.",
         description_ja="絶滅危惧の高地イヌ科。生息数は約500頭。"),

    dict(animal_id="animal_barbary_macaque_025",    name_en="Barbary Macaque",     name_ja="バーバリーマカク",
         category="primate",    rarity_id="rarity_uncommon", base_zoo_value=18,  capture_difficulty=2, growth_rate=3, visitor_appeal=22,
         habitat="mountain",    size="medium", active_time="diurnal",
         description_en="Only macaque native to Africa. Atlas cedar forests.",
         description_ja="アフリカ産唯一のマカク。アトラス山脈のシダー林に棲む。"),

    dict(animal_id="animal_emu_026",                name_en="Emu",                 name_ja="エミュー",
         category="bird",       rarity_id="rarity_uncommon", base_zoo_value=18,  capture_difficulty=2, growth_rate=3, visitor_appeal=24,
         habitat="desert",      size="large", active_time="diurnal",
         description_en="Second-tallest living bird. Curious and fast.",
         description_ja="現生では二番目に背の高い鳥。好奇心旺盛で走るのが速い。"),

    dict(animal_id="animal_bat_eared_fox_027",      name_en="Bat-eared Fox",       name_ja="オオミミギツネ",
         category="small-mammal", rarity_id="rarity_uncommon", base_zoo_value=16, capture_difficulty=2, growth_rate=3, visitor_appeal=26,
         habitat="savanna",     size="small", active_time="cathemeral",
         description_en="Insect-eating fox with cartoon-large ears.",
         description_ja="漫画のような大きな耳を持つ、昆虫を食べるキツネ。"),

    dict(animal_id="animal_marabou_stork_028",      name_en="Marabou Stork",       name_ja="アフリカハゲコウ",
         category="bird",       rarity_id="rarity_uncommon", base_zoo_value=17,  capture_difficulty=2, growth_rate=3, visitor_appeal=15,
         habitat="wetland",     size="large", active_time="diurnal",
         description_en="Massive African scavenger. 3.5-metre wingspan.",
         description_ja="巨大なアフリカのスカベンジャー。翼開長3.5m。"),

    # ------ Rare (10) ----------------------------------------------------
    dict(animal_id="animal_lion_029",               name_en="African Lion",        name_ja="ライオン",
         category="big-cat",    rarity_id="rarity_rare",     base_zoo_value=45,  capture_difficulty=4, growth_rate=2, visitor_appeal=70,
         habitat="savanna",     size="large", active_time="nocturnal",
         description_en="Only truly social big cat. Icon of the savanna.",
         description_ja="真に社会性を持つ唯一の大型ネコ科。サバンナの象徴。"),

    dict(animal_id="animal_black_rhinoceros_030",   name_en="Black Rhinoceros",    name_ja="クロサイ",
         category="hoofed",     rarity_id="rarity_rare",     base_zoo_value=50,  capture_difficulty=4, growth_rate=1, visitor_appeal=55,
         habitat="savanna",     size="huge", active_time="nocturnal",
         description_en="Critically endangered. Pointed prehensile lip for browsing.",
         description_ja="絶滅寸前。伸縮する尖った上唇で葉を食む。"),

    dict(animal_id="animal_chimpanzee_031",         name_en="Chimpanzee",          name_ja="チンパンジー",
         category="primate",    rarity_id="rarity_rare",     base_zoo_value=42,  capture_difficulty=4, growth_rate=2, visitor_appeal=65,
         habitat="rainforest",  size="medium", active_time="diurnal",
         description_en="Closest living relative to humans. Tool-users.",
         description_ja="ヒトに最も近い現生種のひとつ。道具を使う。"),

    dict(animal_id="animal_mountain_gorilla_032",   name_en="Mountain Gorilla",    name_ja="マウンテンゴリラ",
         category="primate",    rarity_id="rarity_rare",     base_zoo_value=55,  capture_difficulty=4, growth_rate=1, visitor_appeal=75,
         habitat="rainforest",  size="huge", active_time="diurnal",
         description_en="Highland gorilla — around 1,000 left in the wild.",
         description_ja="高地に生息するゴリラ。野生では約1,000頭のみ。"),

    dict(animal_id="animal_bongo_033",              name_en="Bongo",               name_ja="ボンゴ",
         category="hoofed",     rarity_id="rarity_rare",     base_zoo_value=38,  capture_difficulty=4, growth_rate=2, visitor_appeal=45,
         habitat="rainforest",  size="large", active_time="nocturnal",
         description_en="Striped forest antelope. Almost never seen in the wild.",
         description_ja="縞模様の森林アンテロープ。野生でその姿はほぼ見られない。"),

    dict(animal_id="animal_jaguar_034",             name_en="Jaguar",              name_ja="ジャガー",
         category="big-cat",    rarity_id="rarity_rare",     base_zoo_value=48,  capture_difficulty=4, growth_rate=2, visitor_appeal=60,
         habitat="rainforest",  size="large", active_time="cathemeral",
         description_en="Third-largest big cat and Amazon apex predator.",
         description_ja="3番目に大きな大型ネコ科。アマゾンの頂点捕食者。"),

    dict(animal_id="animal_giant_anteater_035",     name_en="Giant Anteater",      name_ja="オオアリクイ",
         category="small-mammal", rarity_id="rarity_rare",   base_zoo_value=35,  capture_difficulty=3, growth_rate=2, visitor_appeal=40,
         habitat="rainforest",  size="large", active_time="diurnal",
         description_en="2-metre insect specialist. Distinctive coat pattern.",
         description_ja="全長2mの昆虫食スペシャリスト。特徴的な体色。"),

    dict(animal_id="animal_snow_leopard_036",       name_en="Snow Leopard",        name_ja="ユキヒョウ",
         category="big-cat",    rarity_id="rarity_rare",     base_zoo_value=55,  capture_difficulty=5, growth_rate=2, visitor_appeal=75,
         habitat="mountain",    size="medium", active_time="crepuscular",
         description_en="Ghost of the mountains. Legendary elusiveness in real life.",
         description_ja="山の亡霊。実生活でも伝説的なほど姿を見せない。"),

    dict(animal_id="animal_giant_panda_037",        name_en="Giant Panda",         name_ja="ジャイアントパンダ",
         category="carnivore",  rarity_id="rarity_rare",     base_zoo_value=60,  capture_difficulty=4, growth_rate=1, visitor_appeal=90,
         habitat="mountain",    size="large", active_time="cathemeral",
         description_en="Bamboo-eating bear. Highest visitor draw at Rare tier.",
         description_ja="竹を食べるクマ科。レア帯で最大級の集客力。"),

    dict(animal_id="animal_bengal_tiger_038",       name_en="Bengal Tiger",        name_ja="ベンガルトラ",
         category="big-cat",    rarity_id="rarity_rare",     base_zoo_value=52,  capture_difficulty=4, growth_rate=2, visitor_appeal=72,
         habitat="rainforest",  size="large", active_time="nocturnal",
         description_en="Largest wild cat. Solitary and powerful.",
         description_ja="現生最大の野生ネコ科。単独行動で強靭。"),

    # ------ Epic (7) -----------------------------------------------------
    dict(animal_id="animal_northern_white_rhino_039", name_en="Northern White Rhinoceros", name_ja="キタシロサイ",
         category="hoofed",     rarity_id="rarity_epic",     base_zoo_value=110, capture_difficulty=5, growth_rate=1, visitor_appeal=95,
         habitat="savanna",     size="huge", active_time="diurnal",
         description_en="Functionally extinct. Draft flags special provenance handling.",
         description_ja="事実上絶滅。特殊な由来管理が必要。"),

    dict(animal_id="animal_bonobo_040",             name_en="Bonobo",              name_ja="ボノボ",
         category="primate",    rarity_id="rarity_epic",     base_zoo_value=85,  capture_difficulty=5, growth_rate=2, visitor_appeal=80,
         habitat="rainforest",  size="medium", active_time="diurnal",
         description_en="Peaceful great ape. Only in Congo forests.",
         description_ja="穏やかな大型類人猿。コンゴの森にのみ生息。"),

    dict(animal_id="animal_sumatran_tiger_041",     name_en="Sumatran Tiger",      name_ja="スマトラトラ",
         category="big-cat",    rarity_id="rarity_epic",     base_zoo_value=95,  capture_difficulty=5, growth_rate=2, visitor_appeal=88,
         habitat="rainforest",  size="medium", active_time="nocturnal",
         description_en="Smallest tiger subspecies, less than 400 left.",
         description_ja="最も小型のトラ亜種。400頭未満しか残っていない。"),

    dict(animal_id="animal_sumatran_orangutan_042", name_en="Sumatran Orangutan",  name_ja="スマトラオランウータン",
         category="primate",    rarity_id="rarity_epic",     base_zoo_value=90,  capture_difficulty=4, growth_rate=1, visitor_appeal=82,
         habitat="rainforest",  size="large", active_time="diurnal",
         description_en="Red ape of the canopy. Critically endangered.",
         description_ja="樹冠に棲む赤い類人猿。近絶滅種。"),

    dict(animal_id="animal_polar_bear_043",         name_en="Polar Bear",          name_ja="ホッキョクグマ",
         category="carnivore",  rarity_id="rarity_epic",     base_zoo_value=95,  capture_difficulty=5, growth_rate=2, visitor_appeal=85,
         habitat="tundra",      size="huge", active_time="cathemeral",
         description_en="Ice-dependent apex predator. Long expeditions only.",
         description_ja="氷に依存する頂点捕食者。長時間遠征でのみ狙える。"),

    dict(animal_id="animal_arctic_fox_044",         name_en="Arctic Fox",          name_ja="ホッキョクギツネ",
         category="small-mammal", rarity_id="rarity_epic",   base_zoo_value=75,  capture_difficulty=4, growth_rate=3, visitor_appeal=70,
         habitat="tundra",      size="small", active_time="cathemeral",
         description_en="White in winter, brown in summer. Champion of insulation.",
         description_ja="冬は白、夏は茶。断熱性能の王者。"),

    dict(animal_id="animal_amur_leopard_045",       name_en="Amur Leopard",        name_ja="アムールヒョウ",
         category="big-cat",    rarity_id="rarity_epic",     base_zoo_value=95,  capture_difficulty=5, growth_rate=2, visitor_appeal=85,
         habitat="mountain",    size="medium", active_time="nocturnal",
         description_en="Cold-adapted leopard. Fewer than 120 individuals globally.",
         description_ja="寒冷地適応のヒョウ亜種。世界に120頭未満。"),

    # ------ Legendary (5) ------------------------------------------------
    dict(animal_id="animal_saola_046",              name_en="Saola",               name_ja="サオラ",
         category="hoofed",     rarity_id="rarity_legendary", base_zoo_value=220, capture_difficulty=5, growth_rate=1, visitor_appeal=90,
         habitat="rainforest",  size="medium", active_time="nocturnal",
         description_en="The 'Asian Unicorn'. Almost mythical Vietnamese bovid.",
         description_ja="『アジアの一角獣』。ほぼ神話的なベトナムのウシ科動物。"),

    dict(animal_id="animal_javan_rhinoceros_047",   name_en="Javan Rhinoceros",    name_ja="ジャワサイ",
         category="hoofed",     rarity_id="rarity_legendary", base_zoo_value=260, capture_difficulty=5, growth_rate=1, visitor_appeal=95,
         habitat="rainforest",  size="huge", active_time="cathemeral",
         description_en="Around 76 individuals total. Ujung Kulon National Park only.",
         description_ja="現存総数約76頭。ウジュンクロン国立公園のみ。"),

    dict(animal_id="animal_kakapo_048",             name_en="Kakapo",              name_ja="カカポ",
         category="bird",       rarity_id="rarity_legendary", base_zoo_value=200, capture_difficulty=4, growth_rate=1, visitor_appeal=88,
         habitat="rainforest",  size="medium", active_time="nocturnal",
         description_en="Flightless nocturnal parrot. New Zealand predator-free islands only.",
         description_ja="飛べない夜行性のオウム。ニュージーランドの捕食者のいない島のみ。"),

    dict(animal_id="animal_okapi_049",              name_en="Okapi",               name_ja="オカピ",
         category="hoofed",     rarity_id="rarity_legendary", base_zoo_value=210, capture_difficulty=5, growth_rate=1, visitor_appeal=92,
         habitat="rainforest",  size="large", active_time="diurnal",
         description_en="Congo's forest giraffe. Only discovered by Western science in 1901.",
         description_ja="コンゴの森のキリン。西洋科学に発見されたのは1901年。"),

    dict(animal_id="animal_cross_river_gorilla_050", name_en="Cross River Gorilla", name_ja="クロスリバーゴリラ",
         category="primate",    rarity_id="rarity_legendary", base_zoo_value=240, capture_difficulty=5, growth_rate=1, visitor_appeal=94,
         habitat="rainforest",  size="huge", active_time="diurnal",
         description_en="Rarest great ape. Fewer than 300 individuals. Cameroon–Nigeria border.",
         description_ja="最も希少な類人猿。300頭未満。カメルーンとナイジェリアの国境地帯のみ。"),
]


# =============================================================================
# 4. Hunters  (18 hunters — culturally diverse names, 7 role archetypes)
#
# Rank progression (Draft):  Bronze(1) → Silver(2) → Gold(3) → Platinum(4)
#                            → Diamond(5) → Master(6)
# level = per-Hunter fine-grained tier within their rank (1..10)
#
# capture_bonus / rare_find_bonus / speed_bonus are ADDITIVE percentage
# points applied on top of base ExpeditionRules values. Negative values
# are legal (a beginner may be slower than average).
# =============================================================================

HUNTERS: List[Dict[str, Any]] = [
    # -- Beginner tier (3) --
    dict(hunter_id="hunter_amara_kone_001",  name="Amara Koné",
         rank="Bronze", level=1, specialty="Beginner",
         preferred_biome="savanna",   capture_bonus=0,  rare_find_bonus=-5, speed_bonus=0,
         hire_cost_g=50,
         personality="Cheerful new recruit from the coastal towns.",
         description="Reliable first Hunter for any player. Cheap, competent, no surprises."),

    dict(hunter_id="hunter_kofi_mensah_002", name="Kofi Mensah",
         rank="Bronze", level=2, specialty="Beginner-Forest",
         preferred_biome="rainforest", capture_bonus=0, rare_find_bonus=-5, speed_bonus=0,
         hire_cost_g=60,
         personality="Grew up on the edge of the Congo. Quiet but sure-footed.",
         description="Second-cheapest Hunter. Slight edge in forest biomes."),

    dict(hunter_id="hunter_hana_ito_003",    name="Hana Ito",
         rank="Bronze", level=3, specialty="All-round-Rookie",
         preferred_biome="any",       capture_bonus=+2, rare_find_bonus=0, speed_bonus=0,
         hire_cost_g=90,
         personality="Level-headed rookie from an old expedition family.",
         description="Slightly stronger than the other rookies; often the second Hunter a player unlocks."),

    # -- Rare-find specialists (2) --
    dict(hunter_id="hunter_zara_okafor_004", name="Zara Okafor",
         rank="Silver", level=5, specialty="Rare-Find",
         preferred_biome="savanna",   capture_bonus=0,  rare_find_bonus=+15, speed_bonus=-5,
         hire_cost_g=280,
         personality="Patient tracker. Would rather come back empty-handed than settle for a common find.",
         description="Boosts the odds of returning with a Rare or better, at the cost of speed."),

    dict(hunter_id="hunter_diego_ramirez_005", name="Diego Ramírez",
         rank="Silver", level=6, specialty="Rare-Find-Rainforest",
         preferred_biome="rainforest", capture_bonus=0, rare_find_bonus=+18, speed_bonus=-5,
         hire_cost_g=340,
         personality="Amazon-born biologist. Reads the canopy like a book.",
         description="Rare-find specialist tuned for rainforest maps."),

    # -- Mountain / high-altitude specialists (2) --
    dict(hunter_id="hunter_nadia_kowalski_006", name="Nadia Kowalski",
         rank="Silver", level=5, specialty="Mountain",
         preferred_biome="mountain",  capture_bonus=+10, rare_find_bonus=0, speed_bonus=0,
         hire_cost_g=260,
         personality="Alpine specialist. Speaks softly, climbs faster than anyone.",
         description="Substantially better on mountain maps; average elsewhere."),

    dict(hunter_id="hunter_tenzing_sherpa_007", name="Tenzing Sherpa",
         rank="Gold", level=7, specialty="High-Altitude",
         preferred_biome="mountain",  capture_bonus=+15, rare_find_bonus=+5, speed_bonus=-10,
         hire_cost_g=780,
         personality="Legendary Himalayan guide. The only Hunter fluent above 6000m.",
         description="Only Hunter who reliably returns from extreme mountain expeditions."),

    # -- Nocturnal specialists (2) --
    dict(hunter_id="hunter_yuki_nakamura_008", name="Yuki Nakamura",
         rank="Silver", level=6, specialty="Nocturnal",
         preferred_biome="any",       capture_bonus=+8, rare_find_bonus=+8, speed_bonus=0,
         hire_cost_g=320,
         personality="Sleeps by day, hunts by lantern-light.",
         description="Bonuses apply to nocturnal species regardless of biome (rule TBD)."),

    dict(hunter_id="hunter_aiyana_redcloud_009", name="Aiyana Redcloud",
         rank="Gold", level=7, specialty="Nocturnal-Wilderness",
         preferred_biome="tundra",    capture_bonus=+10, rare_find_bonus=+10, speed_bonus=0,
         hire_cost_g=520,
         personality="Trained in the northern boreal. Reads tracks by moonlight.",
         description="Nocturnal specialist tuned for cold biomes."),

    # -- High-difficulty capture (2) --
    dict(hunter_id="hunter_ravi_chandra_010",  name="Ravi Chandra",
         rank="Gold", level=8, specialty="Hard-Capture-Tropical",
         preferred_biome="rainforest", capture_bonus=+20, rare_find_bonus=0, speed_bonus=0,
         hire_cost_g=680,
         personality="Ex-military tracker. Nothing runs from him twice.",
         description="Straight-up better at succeeding on hard maps; no rare-find bias."),

    dict(hunter_id="hunter_elena_marchetti_011", name="Elena Marchetti",
         rank="Platinum", level=9, specialty="Hard-Capture-Universal",
         preferred_biome="any",       capture_bonus=+25, rare_find_bonus=+5, speed_bonus=-5,
         hire_cost_g=1400,
         personality="Legendary tracker who works in every biome.",
         description="Elite generalist. High cost; near-guaranteed success on high-difficulty maps."),

    # -- High-speed expedition (2) --
    dict(hunter_id="hunter_miguel_santos_012", name="Miguel Santos",
         rank="Silver", level=5, specialty="Speed",
         preferred_biome="any",       capture_bonus=-5, rare_find_bonus=-5, speed_bonus=+30,
         hire_cost_g=240,
         personality="Runs everywhere. Returns before you finish your coffee.",
         description="Best-in-slot for cycling many short expeditions per real-world day."),

    dict(hunter_id="hunter_sara_lindqvist_013", name="Sara Lindqvist",
         rank="Gold", level=7, specialty="Speed-Cold",
         preferred_biome="tundra",    capture_bonus=0,  rare_find_bonus=0,  speed_bonus=+25,
         hire_cost_g=560,
         personality="Skis where others walk. Loves the cold; hates the tropics.",
         description="Speed specialist tuned for tundra maps; slight penalty on hot maps (rule TBD)."),

    # -- Balanced (3) --
    dict(hunter_id="hunter_chen_wei_014",     name="Chen Wei",
         rank="Silver", level=6, specialty="Balanced",
         preferred_biome="any",       capture_bonus=+5, rare_find_bonus=+3, speed_bonus=+3,
         hire_cost_g=300,
         personality="Steady. Reliable. Talks little.",
         description="Balanced upgrade over rookies. First mid-tier Hunter most players buy."),

    dict(hunter_id="hunter_priya_kaur_015",   name="Priya Kaur",
         rank="Gold", level=7, specialty="Balanced-Wetland",
         preferred_biome="wetland",   capture_bonus=+8, rare_find_bonus=+3, speed_bonus=+3,
         hire_cost_g=540,
         personality="Grew up on the Delta. Reads water.",
         description="Balanced Gold-tier Hunter with wetland edge."),

    dict(hunter_id="hunter_kwame_boateng_016", name="Kwame Boateng",
         rank="Gold", level=8, specialty="Balanced-Savanna",
         preferred_biome="savanna",   capture_bonus=+8, rare_find_bonus=+3, speed_bonus=+3,
         hire_cost_g=580,
         personality="Guild instructor turned field Hunter.",
         description="Balanced Gold-tier Hunter with savanna edge."),

    # -- Legendary tier (2, shared scarce) --
    dict(hunter_id="hunter_aiko_fujimori_017", name="Aiko Fujimori",
         rank="Diamond", level=10, specialty="Legendary-All-Round",
         preferred_biome="any",       capture_bonus=+20, rare_find_bonus=+15, speed_bonus=+10,
         hire_cost_g=3800,
         personality="Undefeated in three continents. Books herself.",
         description="Shared scarce Hunter. Only one player may hold her contract at a time."),

    dict(hunter_id="hunter_dr_malik_osei_018", name="Dr. Malik Osei",
         rank="Master", level=10, specialty="Legendary-Rare-Find",
         preferred_biome="rainforest", capture_bonus=+15, rare_find_bonus=+30, speed_bonus=0,
         hire_cost_g=4500,
         personality="Field biologist. Has found the Saola. Once.",
         description="Shared scarce Hunter. Highest rare-find bonus in the game."),
]


# =============================================================================
# 5. MapAnimals  (spawn table)
#
# spawn_weight is a relative draw weight (see Review sheet — the final
# percent probability is not fixed in this Draft).
# capture_modifier is added to the base capture rate for this specific
# Map × Animal pairing (positive = easier here than elsewhere; negative =
# harder here than the species' base).
# minimum_hunter_rank: 0=none, 1=Bronze, 2=Silver, 3=Gold, 4=Platinum,
#                     5=Diamond, 6=Master.
# =============================================================================

def _ma(map_id: str, animal_id: str, spawn_weight: int,
        min_rank: int = 0, capture_mod: int = 0, notes: str = "") -> Dict[str, Any]:
    """Compact constructor for MapAnimal rows."""
    return dict(map_id=map_id, animal_id=animal_id, spawn_weight=spawn_weight,
                minimum_hunter_rank=min_rank, capture_modifier=capture_mod, notes=notes)


MAP_ANIMALS_RAW: List[Dict[str, Any]] = [
    # -- Kenyan Savanna --
    _ma("map_kenyan_savanna_001", "animal_impala_001",            35),
    _ma("map_kenyan_savanna_001", "animal_common_zebra_002",      30),
    _ma("map_kenyan_savanna_001", "animal_blue_wildebeest_003",   28),
    _ma("map_kenyan_savanna_001", "animal_warthog_004",           30),
    _ma("map_kenyan_savanna_001", "animal_cape_buffalo_005",      20),
    _ma("map_kenyan_savanna_001", "animal_chacma_baboon_006",     22),
    _ma("map_kenyan_savanna_001", "animal_vervet_monkey_007",     25),
    _ma("map_kenyan_savanna_001", "animal_aardvark_010",           8, notes="nocturnal — bonus at night"),
    _ma("map_kenyan_savanna_001", "animal_masai_giraffe_015",     15),
    _ma("map_kenyan_savanna_001", "animal_african_elephant_016",  10),
    _ma("map_kenyan_savanna_001", "animal_leopard_017",            8, min_rank=1),
    _ma("map_kenyan_savanna_001", "animal_cheetah_018",            8, min_rank=1),
    _ma("map_kenyan_savanna_001", "animal_spotted_hyena_019",     14),
    _ma("map_kenyan_savanna_001", "animal_lion_029",               6, min_rank=2),
    _ma("map_kenyan_savanna_001", "animal_northern_white_rhino_039", 1, min_rank=5, capture_mod=-20,
        notes="ultra-rare; flagged for special provenance handling"),

    # -- Serengeti Plains --
    _ma("map_serengeti_plains_002", "animal_impala_001",           30),
    _ma("map_serengeti_plains_002", "animal_common_zebra_002",     35),
    _ma("map_serengeti_plains_002", "animal_blue_wildebeest_003",  40, notes="peak during migration season"),
    _ma("map_serengeti_plains_002", "animal_warthog_004",          25),
    _ma("map_serengeti_plains_002", "animal_cape_buffalo_005",     22),
    _ma("map_serengeti_plains_002", "animal_chacma_baboon_006",    18),
    _ma("map_serengeti_plains_002", "animal_masai_giraffe_015",    18),
    _ma("map_serengeti_plains_002", "animal_african_elephant_016", 14),
    _ma("map_serengeti_plains_002", "animal_leopard_017",          10, min_rank=1),
    _ma("map_serengeti_plains_002", "animal_cheetah_018",          14, min_rank=1),
    _ma("map_serengeti_plains_002", "animal_spotted_hyena_019",    18),
    _ma("map_serengeti_plains_002", "animal_african_wild_dog_020",  8, min_rank=2),
    _ma("map_serengeti_plains_002", "animal_lion_029",              9, min_rank=2),
    _ma("map_serengeti_plains_002", "animal_black_rhinoceros_030",  3, min_rank=3),

    # -- Okavango Delta --
    _ma("map_okavango_delta_003",   "animal_cape_buffalo_005",     22),
    _ma("map_okavango_delta_003",   "animal_hippopotamus_011",     28),
    _ma("map_okavango_delta_003",   "animal_nile_crocodile_012",   26),
    _ma("map_okavango_delta_003",   "animal_african_elephant_016", 16),
    _ma("map_okavango_delta_003",   "animal_serval_021",           14, min_rank=1),
    _ma("map_okavango_delta_003",   "animal_african_wild_dog_020",  9, min_rank=2),
    _ma("map_okavango_delta_003",   "animal_marabou_stork_028",    18),

    # -- Namib Desert --
    _ma("map_namib_desert_004",     "animal_springbok_008",        28),
    _ma("map_namib_desert_004",     "animal_meerkat_009",          32),
    _ma("map_namib_desert_004",     "animal_aardvark_010",         10, notes="nocturnal"),
    _ma("map_namib_desert_004",     "animal_fennec_fox_013",       25),
    _ma("map_namib_desert_004",     "animal_bat_eared_fox_027",    18, min_rank=1),

    # -- Atlas Mountains --
    _ma("map_atlas_mountains_005",  "animal_barbary_macaque_025",  28, min_rank=1),
    _ma("map_atlas_mountains_005",  "animal_ethiopian_wolf_024",    9, min_rank=2, notes="highland canid, rare here"),
    _ma("map_atlas_mountains_005",  "animal_caracal_022",          18, min_rank=1),
    _ma("map_atlas_mountains_005",  "animal_fennec_fox_013",       14),
    _ma("map_atlas_mountains_005",  "animal_leopard_017",           7, min_rank=2, capture_mod=-5,
        notes="Barbary leopard — extremely rare in reality; flag for review"),

    # -- Kilimanjaro Slopes --
    _ma("map_kilimanjaro_slopes_006","animal_ethiopian_wolf_024",  10, min_rank=2),
    _ma("map_kilimanjaro_slopes_006","animal_leopard_017",         12, min_rank=2),
    _ma("map_kilimanjaro_slopes_006","animal_serval_021",          10, min_rank=1),
    _ma("map_kilimanjaro_slopes_006","animal_mountain_gorilla_032",  6, min_rank=3, capture_mod=-10,
        notes="Kilimanjaro region — placement is a game abstraction, revisit for realism"),
    _ma("map_kilimanjaro_slopes_006","animal_bongo_033",            5, min_rank=3),
    _ma("map_kilimanjaro_slopes_006","animal_chacma_baboon_006",   15),

    # -- Congo Rainforest --
    _ma("map_congo_rainforest_007", "animal_leopard_017",          10, min_rank=2),
    _ma("map_congo_rainforest_007", "animal_chimpanzee_031",       14, min_rank=2),
    _ma("map_congo_rainforest_007", "animal_mountain_gorilla_032",  8, min_rank=3),
    _ma("map_congo_rainforest_007", "animal_bongo_033",            10, min_rank=3),
    _ma("map_congo_rainforest_007", "animal_bonobo_040",            5, min_rank=4, notes="Congo Basin south of the river"),
    _ma("map_congo_rainforest_007", "animal_okapi_049",             2, min_rank=5, capture_mod=-15),
    _ma("map_congo_rainforest_007", "animal_cross_river_gorilla_050", 1, min_rank=6, capture_mod=-25,
        notes="Cross River drainage — game abstraction that includes it in Congo pool"),

    # -- Amazon Rainforest --
    _ma("map_amazon_rainforest_008", "animal_jaguar_034",          10, min_rank=2),
    _ma("map_amazon_rainforest_008", "animal_giant_anteater_035",  16, min_rank=1),

    # -- Borneo Jungle --
    _ma("map_borneo_jungle_009",     "animal_leopard_017",         12, min_rank=2, notes="clouded leopard as placeholder — revisit"),
    _ma("map_borneo_jungle_009",     "animal_bengal_tiger_038",     8, min_rank=3),
    _ma("map_borneo_jungle_009",     "animal_sumatran_tiger_041",   4, min_rank=4),
    _ma("map_borneo_jungle_009",     "animal_sumatran_orangutan_042", 6, min_rank=4),
    _ma("map_borneo_jungle_009",     "animal_saola_046",            1, min_rank=6, capture_mod=-30,
        notes="Vietnam/Laos in reality; game abstraction places it in the Southeast Asian pool"),
    _ma("map_borneo_jungle_009",     "animal_javan_rhinoceros_047", 1, min_rank=6, capture_mod=-30),
    _ma("map_borneo_jungle_009",     "animal_kakapo_048",           1, min_rank=6, capture_mod=-30,
        notes="New Zealand endemic in reality; placement flagged for review"),

    # -- Australian Outback --
    _ma("map_australian_outback_010", "animal_red_kangaroo_023",   32, min_rank=1),
    _ma("map_australian_outback_010", "animal_dingo_014",          28),
    _ma("map_australian_outback_010", "animal_emu_026",            26, min_rank=1),

    # -- Himalayan Foothills --
    _ma("map_himalayan_foothills_011", "animal_snow_leopard_036",  10, min_rank=3),
    _ma("map_himalayan_foothills_011", "animal_giant_panda_037",    7, min_rank=3, notes="Sichuan cordillera game abstraction"),
    _ma("map_himalayan_foothills_011", "animal_amur_leopard_045",   3, min_rank=4, capture_mod=-15,
        notes="Amur is Russian Far East; game abstraction pools it under 'Asian mountain'"),

    # -- Arctic Tundra --
    _ma("map_arctic_tundra_012",     "animal_polar_bear_043",       8, min_rank=4),
    _ma("map_arctic_tundra_012",     "animal_arctic_fox_044",      16, min_rank=3),
]


# =============================================================================
# 6. HunterSkills  (data-dictionary sheet)
#
# The Hunters sheet embeds capture_bonus / rare_find_bonus / speed_bonus
# columns directly (see report §"design decision"). This sheet documents
# what those columns mean and enumerates the skill catalogue for future
# expansion.
# =============================================================================

HUNTER_SKILLS: List[Dict[str, Any]] = [
    dict(skill_id="skill_capture_bonus",
         name_en="Capture Bonus",     name_ja="捕獲補正",
         effect_type="additive_percent",
         effect_min=-20, effect_max=+30,
         description="Added to the base capture success rate. Applied on top of Map difficulty."),

    dict(skill_id="skill_rare_find_bonus",
         name_en="Rare Find Bonus",   name_ja="レア発見補正",
         effect_type="additive_percent",
         effect_min=-20, effect_max=+30,
         description="Bias toward drawing a rarer Animal from the Map's spawn table when a capture succeeds."),

    dict(skill_id="skill_speed_bonus",
         name_en="Speed Bonus",       name_ja="速度補正",
         effect_type="additive_percent",
         effect_min=-20, effect_max=+30,
         description="Reduces the expedition_minutes for the assigned Map by this percentage."),

    dict(skill_id="skill_biome_affinity",
         name_en="Biome Affinity",    name_ja="バイオーム適性",
         effect_type="biome_multiplier",
         effect_min=0, effect_max=0,
         description="A Hunter's preferred_biome grants additional bonuses when that biome matches the Map. "
                     "The exact bonus size is a Draft parameter (see ExpeditionRules)."),
]


# =============================================================================
# 7. ExpeditionRules  (numeric knobs — all Draft)
# =============================================================================

EXPEDITION_RULES: List[Dict[str, Any]] = [
    dict(rule_id="expedition_rule_base_success_rate",
         rule_name="Base success rate",
         value=60, unit="percent",
         description="Baseline chance an expedition returns with any capture, before Map difficulty or Hunter skill."),

    dict(rule_id="expedition_rule_minimum_minutes",
         rule_name="Minimum expedition minutes",
         value=5, unit="minutes",
         description="Absolute floor after speed bonuses. A speed-Hunter cannot make a Map instant."),

    dict(rule_id="expedition_rule_maximum_minutes",
         rule_name="Maximum expedition minutes",
         value=1440, unit="minutes",
         description="Absolute ceiling (24h). Applies even before slower Hunters penalise a long Map further."),

    dict(rule_id="expedition_rule_base_capture_attempts",
         rule_name="Base capture attempts per expedition",
         value=1, unit="count",
         description="Draft: single-attempt per expedition. Multi-attempt logic is deferred."),

    dict(rule_id="expedition_rule_failure_penalty_g",
         rule_name="Failure penalty (G)",
         value=0, unit="G",
         description="Draft: no explicit G penalty on failure beyond the sunk dispatch cost. See Review."),

    dict(rule_id="expedition_rule_release_reward_ratio",
         rule_name="Release reward ratio",
         value=0, unit="ratio",
         description="Draft: Releasing an Animal returns no G. See Review — human decision needed."),

    dict(rule_id="expedition_rule_biome_affinity_bonus",
         rule_name="Biome affinity bonus",
         value=10, unit="percent",
         description="Additive capture bonus when Hunter.preferred_biome matches Map.biome."),

    dict(rule_id="expedition_rule_biome_mismatch_penalty",
         rule_name="Biome mismatch penalty",
         value=0, unit="percent",
         description="Draft: no penalty for mismatch. Purely reward the match. See Review."),

    dict(rule_id="expedition_rule_hunter_rank_gate_soft",
         rule_name="Hunter rank gate mode",
         value=1, unit="mode (1=hard, 0=soft)",
         description="Draft: hard gate. A too-low-rank Hunter cannot be dispatched to a Map. Soft mode would allow it with a heavy penalty."),

    dict(rule_id="expedition_rule_rare_find_source",
         rule_name="Rare Find source",
         value=1, unit="mode (1=post-capture, 0=pre-capture)",
         description="Draft: after a successful capture, the game rolls the actual species from the Map's spawn table weighted by rarity + rare_find_bonus."),
]


# =============================================================================
# 8. Review  (points the human is asked to resolve on the Excel itself)
# =============================================================================

REVIEW: List[Dict[str, Any]] = [
    dict(review_id="review_001", category="Scope",
         question="12 maps + 50 animals + 18 hunters — is the ratio right for v0.1?",
         current_proposal="12 / 50 / 18",
         reason="Matches ADR-0002 §3 (design target 120 species) at ~40% coverage. Enough for early gameplay tuning without over-designing.",
         priority="high", decision="", notes=""),

    dict(review_id="review_002", category="Scope",
         question="Are 5 rarity tiers (Common → Legendary) enough?",
         current_proposal="5 tiers, no Mythic.",
         reason="ADR-0002 §5 references 'rare traits' + Species-rarity + individual-rarity as separate; adding a 6th tier before those interact is premature.",
         priority="high", decision="", notes=""),

    dict(review_id="review_003", category="Regions",
         question="7 of 12 maps are African; is that the right balance vs global variety?",
         current_proposal="7 African, 2 South American / Southeast Asian, 1 Australian, 1 South Asian, 1 Arctic.",
         reason="ADR-0002 §8 states 'An African-inspired setting is the working assumption', softened by the need for a diverse species pool.",
         priority="high", decision="", notes=""),

    dict(review_id="review_004", category="Regions",
         question="Should some legendary/epic Animals be temporarily placed outside their real range for gameplay (e.g. Kakapo in Borneo pool)?",
         current_proposal="Yes, with a 'game abstraction' note per row.",
         reason="Without a NZ / island map, several Legendary species have no home. Options: (a) accept abstraction (current), (b) add more maps, (c) drop those species from v0.1.",
         priority="high", decision="", notes=""),

    dict(review_id="review_005", category="Hunters",
         question="6-step rank system (Bronze → Master) — right granularity?",
         current_proposal="Bronze, Silver, Gold, Platinum, Diamond, Master + per-Hunter level 1..10.",
         reason="Enough steps to feel like progress; not so many that Guild UI becomes cluttered.",
         priority="medium", decision="", notes=""),

    dict(review_id="review_006", category="Hunters",
         question="Is 18 Hunters the right initial roster size?",
         current_proposal="18 (3 beginner / 2 rare-find / 2 mountain / 2 nocturnal / 2 hard / 2 speed / 3 balanced / 2 legendary).",
         reason="Each archetype has ≥1 example; two Legendary reflects ADR-0002 §7 'shared scarce Hunters'.",
         priority="medium", decision="", notes=""),

    dict(review_id="review_007", category="Hunters",
         question="Represent hunter skills as columns on Hunters sheet (current) or as a HunterSkillAssignments join sheet?",
         current_proposal="Columns on Hunters sheet + HunterSkills sheet as data dictionary.",
         reason="Draft simplicity. Every Hunter has the same 3 numeric skills today; a join sheet would double the review surface with no information gain. Switch to join sheet the moment a Hunter can have variable skill sets.",
         priority="medium", decision="", notes=""),

    dict(review_id="review_008", category="Maps",
         question="Map unlock model — is unlock_rank (minimum Hunter rank) the right axis?",
         current_proposal="Yes. Unlock by owning a Hunter of the required rank rather than by Player level or story progress.",
         reason="Progression is Hunter-driven per ADR-0002 §1. Player-level or story-based unlocks were not decided in ADR-0002.",
         priority="high", decision="", notes=""),

    dict(review_id="review_009", category="Expedition",
         question="Expedition minutes — do the 10-min → 1200-min extremes match the ADR-0002 §9 range (~10 min to ~24 h)?",
         current_proposal="Kenyan Savanna 10 min, Arctic Tundra 1200 min (20 h). Within the ADR range.",
         reason="Fits the 'session-length variety' intent. Extreme maps are locked behind rank 4+, so they don't ambush new players.",
         priority="medium", decision="", notes=""),

    dict(review_id="review_010", category="Economy",
         question="G cost levels — dispatch 50G (starter) to 2400G (Arctic). Reasonable early-game gradient?",
         current_proposal="~48× spread between cheapest and most expensive dispatch.",
         reason="Draft. Actual G income formula not yet decided (ADR-0002 §13); numbers are placeholders for shape, not magnitude.",
         priority="high", decision="", notes=""),

    dict(review_id="review_011", category="Economy",
         question="Animal base_zoo_value — Common ~10 → Legendary base ~200-260. Scale OK?",
         current_proposal="Legendary is 20-26× a Common.",
         reason="Meaningfully rewarding but not so extreme that a single Legendary trivialises the Zoo Value ranking (ADR-0002 §12 diminishing returns).",
         priority="high", decision="", notes=""),

    dict(review_id="review_012", category="Presentation",
         question="How should visitor_appeal be used?",
         current_proposal="Feeds a per-Animal Visitor contribution in the future Visitor formula (ADR-0002 §13).",
         reason="Not consumed by v0.1 gameplay; recorded so the balance conversation can happen.",
         priority="medium", decision="", notes=""),

    dict(review_id="review_013", category="Gameplay",
         question="What happens on capture failure — beyond G sunk cost?",
         current_proposal="No additional penalty (ExpeditionRules.failure_penalty_g = 0).",
         reason="ADR-0002 §10 defines only 'capture success or no capture'; extra penalty is a design choice.",
         priority="medium", decision="", notes=""),

    dict(review_id="review_014", category="Gameplay",
         question="When the player Releases an Animal, do they get any G back?",
         current_proposal="No (release_reward_ratio = 0).",
         reason="ADR-0002 §11 explicitly rejects animal-selling as a core loop; Release is disposal. A tiny 'conservation refund' is possible but not decided.",
         priority="high", decision="", notes=""),

    dict(review_id="review_015", category="Guild",
         question="Hunter hire model — one-off contract cost only? Or ongoing upkeep?",
         current_proposal="One-off cost per contract (hire_cost_g). Contract duration TBD.",
         reason="ADR-0002 §7 leaves contract-duration open. Upkeep would be an entirely new mechanic; defer.",
         priority="high", decision="", notes=""),

    dict(review_id="review_016", category="Balance",
         question="Northern White Rhinoceros in v0.1 — appropriate?",
         current_proposal="Included in Epic with a note flagging real-world 'functionally extinct' status.",
         reason="Powerful narrative moment if captured, but real animals are down to 2 individuals in reality (2026). Might feel disrespectful. Human call.",
         priority="high", decision="", notes=""),

    dict(review_id="review_017", category="Documentation",
         question="Is a bilingual (JA/EN) column pair the right multilingual approach?",
         current_proposal="Yes — every player-visible name has name_en + name_ja columns.",
         reason="Matches the docs/reports bilingual policy. Adding more languages later is additive.",
         priority="low", decision="", notes=""),
]


# =============================================================================
# ID naming rules (for the README / self-check)
# =============================================================================

ID_PREFIXES = {
    "map":            "map_<slug>_<###>",
    "animal":         "animal_<slug>_<###>",
    "hunter":         "hunter_<given>_<family>_<###>",
    "rarity":         "rarity_<name>",
    "hunter_skill":   "skill_<name>",
    "expedition_rule":"expedition_rule_<name>",
    "review":         "review_<###>",
}


# =============================================================================
# Excel writing
# =============================================================================

HEADER_FILL = PatternFill(start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFFFF")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _write_sheet(ws, header: List[str], rows: List[Dict[str, Any]], widths: Dict[str, int]) -> None:
    """Write a sheet with header row, freeze pane, filter, column widths."""
    # Header row
    for col_idx, key in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx, value=key)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, key in enumerate(header, start=1):
            value = row.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = LEFT

    # Column widths
    for col_idx, key in enumerate(header, start=1):
        width = widths.get(key, 18)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze the header
    ws.freeze_panes = "A2"

    # Autofilter over the header row
    last_col_letter = get_column_letter(len(header))
    ws.auto_filter.ref = f"A1:{last_col_letter}{max(len(rows) + 1, 2)}"


# =============================================================================
# Validation
# =============================================================================

def validate() -> Tuple[bool, List[str]]:
    problems: List[str] = []

    # Unique IDs per sheet
    for name, rows, key in [
        ("Rarities",         RARITIES,         "rarity_id"),
        ("Maps",             MAPS,             "map_id"),
        ("Animals",          ANIMALS,          "animal_id"),
        ("Hunters",          HUNTERS,          "hunter_id"),
        ("HunterSkills",     HUNTER_SKILLS,    "skill_id"),
        ("ExpeditionRules",  EXPEDITION_RULES, "rule_id"),
        ("Review",           REVIEW,           "review_id"),
    ]:
        seen = set()
        for r in rows:
            v = r[key]
            if v in seen:
                problems.append(f"{name}: duplicate {key} = {v!r}")
            seen.add(v)

    # FK: MapAnimals.map_id → Maps.map_id, animal_id → Animals.animal_id
    map_ids = {m["map_id"] for m in MAPS}
    animal_ids = {a["animal_id"] for a in ANIMALS}
    rarity_ids = {r["rarity_id"] for r in RARITIES}

    for ma in MAP_ANIMALS_RAW:
        if ma["map_id"] not in map_ids:
            problems.append(f"MapAnimals: map_id {ma['map_id']!r} not in Maps")
        if ma["animal_id"] not in animal_ids:
            problems.append(f"MapAnimals: animal_id {ma['animal_id']!r} not in Animals")

    # FK: Animals.rarity_id → Rarities.rarity_id
    for a in ANIMALS:
        if a["rarity_id"] not in rarity_ids:
            problems.append(f"Animals: rarity_id {a['rarity_id']!r} not in Rarities (animal {a['animal_id']})")

    # Non-null / non-empty required fields
    def _require(name: str, rows: List[Dict[str, Any]], fields: List[str]) -> None:
        for r in rows:
            for f in fields:
                v = r.get(f)
                if v is None or (isinstance(v, str) and not v.strip()):
                    problems.append(f"{name}: {r.get('name_en') or r.get('rule_id') or r.get('review_id') or '?'} — empty {f}")

    _require("Rarities",        RARITIES,        ["rarity_id", "name_en", "name_ja", "sort_order", "base_multiplier"])
    _require("Maps",            MAPS,            ["map_id", "name_en", "name_ja", "region", "biome",
                                                  "difficulty", "expedition_minutes", "base_cost_g"])
    _require("Animals",         ANIMALS,         ["animal_id", "name_en", "name_ja", "rarity_id",
                                                  "base_zoo_value", "capture_difficulty", "habitat"])
    _require("Hunters",         HUNTERS,         ["hunter_id", "name", "rank", "specialty",
                                                  "preferred_biome", "hire_cost_g"])
    _require("HunterSkills",    HUNTER_SKILLS,   ["skill_id", "name_en", "effect_type"])
    _require("ExpeditionRules", EXPEDITION_RULES,["rule_id", "rule_name", "value", "unit"])
    _require("Review",          REVIEW,          ["review_id", "category", "question", "current_proposal", "priority"])

    # Numeric range sanity
    for m in MAPS:
        if not (1 <= m["difficulty"] <= 5):
            problems.append(f"Maps.{m['map_id']}: difficulty out of range: {m['difficulty']}")
        if not (0 <= m["unlock_rank"] <= 6):
            problems.append(f"Maps.{m['map_id']}: unlock_rank out of range: {m['unlock_rank']}")
    for a in ANIMALS:
        if not (1 <= a["capture_difficulty"] <= 5):
            problems.append(f"Animals.{a['animal_id']}: capture_difficulty out of range")
        if not (1 <= a["visitor_appeal"] <= 100):
            problems.append(f"Animals.{a['animal_id']}: visitor_appeal out of 1..100")

    # Every animal used somewhere on MapAnimals? Not required, but flag.
    used = {ma["animal_id"] for ma in MAP_ANIMALS_RAW}
    for a in ANIMALS:
        if a["animal_id"] not in used:
            problems.append(f"NOTICE: Animal {a['animal_id']!r} has no MapAnimals row")

    return (len([p for p in problems if not p.startswith("NOTICE")]) == 0, problems)


# =============================================================================
# Main
# =============================================================================

def main() -> int:
    ok, problems = validate()
    for p in problems:
        print(("[!] " if not p.startswith("NOTICE") else "[i] ") + p)
    if not ok:
        print("\nvalidation failed — refusing to write xlsx")
        return 1

    wb = Workbook()
    # openpyxl creates one default sheet — rename to Rarities and reuse
    ws0 = wb.active
    ws0.title = "Rarities"
    _write_sheet(ws0,
                 header=["rarity_id", "name_en", "name_ja", "sort_order",
                         "base_multiplier", "description"],
                 rows=RARITIES,
                 widths={"rarity_id": 22, "name_en": 14, "name_ja": 14,
                         "sort_order": 12, "base_multiplier": 16,
                         "description": 55})

    _write_sheet(wb.create_sheet("Maps"),
                 header=["map_id", "name_ja", "name_en", "region", "biome",
                         "difficulty", "unlock_rank", "expedition_minutes",
                         "base_cost_g", "risk_level",
                         "description_ja", "description_en"],
                 rows=MAPS,
                 widths={"map_id": 32, "name_ja": 22, "name_en": 22, "region": 16,
                         "biome": 14, "difficulty": 11, "unlock_rank": 12,
                         "expedition_minutes": 20, "base_cost_g": 14, "risk_level": 12,
                         "description_ja": 55, "description_en": 55})

    _write_sheet(wb.create_sheet("Animals"),
                 header=["animal_id", "name_ja", "name_en", "category",
                         "rarity_id", "base_zoo_value", "capture_difficulty",
                         "growth_rate", "visitor_appeal",
                         "habitat", "size", "active_time",
                         "description_ja", "description_en"],
                 rows=ANIMALS,
                 widths={"animal_id": 34, "name_ja": 20, "name_en": 22, "category": 14,
                         "rarity_id": 18, "base_zoo_value": 14, "capture_difficulty": 18,
                         "growth_rate": 12, "visitor_appeal": 14,
                         "habitat": 14, "size": 10, "active_time": 12,
                         "description_ja": 55, "description_en": 55})

    _write_sheet(wb.create_sheet("Hunters"),
                 header=["hunter_id", "name", "rank", "level", "specialty",
                         "preferred_biome", "capture_bonus", "rare_find_bonus",
                         "speed_bonus", "hire_cost_g", "personality", "description"],
                 rows=HUNTERS,
                 widths={"hunter_id": 32, "name": 22, "rank": 12, "level": 8,
                         "specialty": 24, "preferred_biome": 16,
                         "capture_bonus": 14, "rare_find_bonus": 16,
                         "speed_bonus": 12, "hire_cost_g": 12,
                         "personality": 45, "description": 55})

    # Add compact map_animal_id at write time
    map_animal_rows = []
    for i, ma in enumerate(MAP_ANIMALS_RAW, start=1):
        row = {"map_animal_id": f"map_animal_{i:03d}"}
        row.update(ma)
        map_animal_rows.append(row)
    _write_sheet(wb.create_sheet("MapAnimals"),
                 header=["map_animal_id", "map_id", "animal_id", "spawn_weight",
                         "minimum_hunter_rank", "capture_modifier", "notes"],
                 rows=map_animal_rows,
                 widths={"map_animal_id": 18, "map_id": 32, "animal_id": 34,
                         "spawn_weight": 14, "minimum_hunter_rank": 20,
                         "capture_modifier": 16, "notes": 55})

    _write_sheet(wb.create_sheet("HunterSkills"),
                 header=["skill_id", "name_ja", "name_en", "effect_type",
                         "effect_min", "effect_max", "description"],
                 rows=HUNTER_SKILLS,
                 widths={"skill_id": 24, "name_ja": 20, "name_en": 22,
                         "effect_type": 22, "effect_min": 12, "effect_max": 12,
                         "description": 60})

    _write_sheet(wb.create_sheet("ExpeditionRules"),
                 header=["rule_id", "rule_name", "value", "unit", "description"],
                 rows=EXPEDITION_RULES,
                 widths={"rule_id": 40, "rule_name": 30, "value": 10,
                         "unit": 24, "description": 65})

    _write_sheet(wb.create_sheet("Review"),
                 header=["review_id", "category", "question", "current_proposal",
                         "reason", "priority", "decision", "notes"],
                 rows=REVIEW,
                 widths={"review_id": 14, "category": 14, "question": 50,
                         "current_proposal": 40, "reason": 60,
                         "priority": 10, "decision": 20, "notes": 20})

    out = Path(__file__).parent / "WildLive-Game-Master-Draft-v0.1.xlsx"
    wb.save(out)
    print(f"wrote {out}")
    print(f"    sheets:            {len(wb.sheetnames)}  ({', '.join(wb.sheetnames)})")
    print(f"    Rarities:          {len(RARITIES)}")
    print(f"    Maps:              {len(MAPS)}")
    print(f"    Animals:           {len(ANIMALS)}")
    print(f"    Hunters:           {len(HUNTERS)}")
    print(f"    MapAnimals:        {len(MAP_ANIMALS_RAW)}")
    print(f"    HunterSkills:      {len(HUNTER_SKILLS)}")
    print(f"    ExpeditionRules:   {len(EXPEDITION_RULES)}")
    print(f"    Review items:      {len(REVIEW)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
